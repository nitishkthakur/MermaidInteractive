#!/usr/bin/env python3
"""
excel_lineage.py
================
Read an Excel workbook (.xlsx / .xlsm) and build a sheet-level dependency
diagram by parsing cell formulas.

Each *node* in the diagram is one worksheet.  An edge  A ──► B  means
"sheet B contains at least one formula that reads from sheet A."
1000 formulas between the same two sheets still produce a single edge.

Upstream workbook references (e.g. ='[Upstream.xlsx]Data'!A1) are detected
and shown as separate subgraphs.

Usage
-----
    python excel_lineage.py  myfile.xlsx
    python excel_lineage.py  myfile.xlsx  -o lineage.html
"""

import re
import os
import sys
import html as html_module
import argparse
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("openpyxl is required:  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Import the proven HTML generator from the parent package
_THIS_DIR   = os.path.dirname(os.path.abspath(__file__))
_PARENT_DIR = os.path.dirname(_THIS_DIR)
if _PARENT_DIR not in sys.path:
    sys.path.insert(0, _PARENT_DIR)
from mermaid_interactive import generate_html as _mi_generate_html  # noqa: E402


# ---------------------------------------------------------------------------
# Formula reference extraction
# ---------------------------------------------------------------------------

# Matches every sheet-reference token inside a formula string.
# Priority (first alternative wins):
#   1. External quoted:   '[Path/Book.xlsx]Sheet Name'!
#   2. External unquoted: [Book.xlsx]SheetName!          (no spaces in names)
#   3. Internal quoted:   'Sheet Name'!
#   4. Internal unquoted: SheetName!
#
# Named groups:
#   ext_q_wb / ext_q_sh  – external, quoted
#   ext_u_wb / ext_u_sh  – external, unquoted
#   int_q                – internal, quoted
#   int_u                – internal, unquoted
_SHEET_REF_RE = re.compile(
    r"""
    (?:
      # ── External quoted: '[path/Book.xlsx]Sheet Name'! ──────────────────
      '\[(?P<ext_q_wb>[^\[\]']+)\](?P<ext_q_sh>[^']*)'\s*!
      |
      # ── External unquoted: [Book.xlsx]SheetName! ────────────────────────
      \[(?P<ext_u_wb>[^\[\]']+)\](?P<ext_u_sh>[A-Za-z0-9_.][A-Za-z0-9_. ]*)\s*!
      |
      # ── Internal quoted: 'Sheet Name'! ──────────────────────────────────
      '(?P<int_q>[^'\[\]]+)'\s*!
      |
      # ── Internal unquoted: SheetName! ────────────────────────────────────
      # Negative lookbehind on ] prevents re-matching an ext-unquoted sheet.
      (?<!\])(?P<int_u>[A-Za-z_][A-Za-z0-9_]*)\s*!
    )
    """,
    re.IGNORECASE | re.VERBOSE,
)


def extract_refs(formula: str) -> list[tuple[str | None, str]]:
    """
    Return all sheet references found in *formula*.

    Each element is ``(workbook_filename_or_None, sheet_name)``.
    ``None`` means a same-workbook reference.

    Parameters
    ----------
    formula : str
        Raw cell formula string (must start with ``=``).

    Returns
    -------
    list of (workbook, sheet) tuples — deduplicated within this call.
    """
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return []

    # Strip double-quoted string literals so we don't match sheet-name-like
    # text inside strings (e.g. INDIRECT("Sheet1!A1") must not produce a ref).
    formula_stripped = re.sub(r'"[^"]*"', '""', formula)

    seen: set[tuple[str | None, str]] = set()
    results: list[tuple[str | None, str]] = []

    for m in _SHEET_REF_RE.finditer(formula_stripped):
        if m.group("ext_q_wb") is not None:
            wb = _basename(m.group("ext_q_wb").strip())
            sh = m.group("ext_q_sh").strip()
        elif m.group("ext_u_wb") is not None:
            wb = _basename(m.group("ext_u_wb").strip())
            sh = m.group("ext_u_sh").strip()
        elif m.group("int_q") is not None:
            wb = None
            sh = m.group("int_q").strip()
        else:
            wb = None
            sh = m.group("int_u").strip()

        if sh and (wb, sh) not in seen:
            seen.add((wb, sh))
            results.append((wb, sh))

    return results


def _basename(path: str) -> str:
    """Return just the filename from any absolute or relative path."""
    # Normalise Windows back-slashes before splitting
    return os.path.basename(path.replace("\\", "/"))


# ---------------------------------------------------------------------------
# Graph data model
# ---------------------------------------------------------------------------

class SheetNode:
    """
    Represents one worksheet, possibly in an external workbook.

    Attributes
    ----------
    workbook : str or None
        Filename of the workbook (e.g. ``"Upstream.xlsx"``).
        ``None`` means the currently-analysed workbook.
    sheet : str
        Worksheet name (e.g. ``"Sheet1"``).
    """

    __slots__ = ("workbook", "sheet")

    def __init__(self, workbook: str | None, sheet: str) -> None:
        self.workbook = workbook
        self.sheet = sheet

    # ── Mermaid-safe node ID ──────────────────────────────────────────────

    @property
    def node_id(self) -> str:
        """
        A Mermaid-safe node identifier.

        Format: ``CUR__<sheet>`` for internal sheets,
                ``EXT__<workbook>__<sheet>`` for external ones.
        All non-alphanumeric characters are replaced with ``_``.
        """
        if self.workbook:
            raw = f"EXT__{self.workbook}__{self.sheet}"
        else:
            raw = f"CUR__{self.sheet}"
        safe = re.sub(r"[^A-Za-z0-9_]", "_", raw)
        # Mermaid node IDs must start with a letter
        if safe and not safe[0].isalpha():
            safe = "N_" + safe
        return safe

    @property
    def label(self) -> str:
        """Human-readable label shown in the diagram node."""
        return self.sheet

    # ── Python special methods ────────────────────────────────────────────

    def __eq__(self, other: object) -> bool:
        return (
            isinstance(other, SheetNode)
            and self.workbook == other.workbook
            and self.sheet == other.sheet
        )

    def __hash__(self) -> int:
        return hash((self.workbook, self.sheet))

    def __repr__(self) -> str:
        return f"SheetNode({self.workbook!r}, {self.sheet!r})"


# ---------------------------------------------------------------------------
# External connections (ODBC / Power Query / Web)
# ---------------------------------------------------------------------------

# XML namespaces used in Office Open XML
_OOXML_NS  = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_X15_NS    = "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"
_PQ_EXT_URI = "{DE250136-89BD-433C-8126-D09CA5730AF9}"  # Power Query ext URI

# Connection type integers defined by the OOXML spec
_TYPE_ODBC   = 1
_TYPE_DAO    = 2
_TYPE_WEB    = 4
_TYPE_OLEDB  = 5


class ExternalConnectionNode:
    """
    Represents an upstream data source connected to the workbook via ODBC,
    Power Query, or a web query.

    Attributes
    ----------
    conn_type : str
        ``"odbc"``, ``"powerquery"``, ``"web"``, or ``"oledb"``
    name : str
        Connection / query name as stored in Excel.
    source : str
        Meaningful source identifier: DSN, server, URL, or query name.
    """

    __slots__ = ("conn_type", "name", "source", "wb_scope")

    def __init__(
        self,
        conn_type: str,
        name: str,
        source: str,
        wb_scope: str = "",
    ) -> None:
        self.conn_type = conn_type
        self.name      = name
        self.source    = source
        # wb_scope disambiguates connections that live in an upstream workbook.
        # Leave empty ("") for connections of the currently-analysed file.
        self.wb_scope  = wb_scope

    @property
    def node_id(self) -> str:
        prefix = {
            "odbc":       "ODBC",
            "powerquery": "PQ",
            "web":        "WEB",
            "oledb":      "OLEDB",
        }.get(self.conn_type, "EXT")
        # Clean source/name; prefix workbook scope when set so IDs are unique
        # across workbooks even when two files share the same connection name.
        raw = re.sub(r"[^A-Za-z0-9_]", "_", self.source or self.name)
        raw = re.sub(r"_+", "_", raw).strip("_") or "unnamed"
        if self.wb_scope:
            scope = re.sub(r"[^A-Za-z0-9_]", "_", self.wb_scope)
            scope = re.sub(r"_+", "_", scope).strip("_")
            safe  = f"{prefix}_{scope}__{raw}"
        else:
            safe = f"{prefix}__{raw}"
        if safe and not safe[0].isalpha():
            safe = "N_" + safe
        return safe[:80]

    @property
    def label(self) -> str:
        if self.source and self.source != self.name:
            return f"{self.name}\\n({self.source})"
        return self.name

    @property
    def mermaid_class(self) -> str:
        return {
            "odbc":       "odbcNode",
            "powerquery": "pqNode",
            "web":        "webNode",
            "oledb":      "oledbNode",
        }.get(self.conn_type, "extNode")

    def __eq__(self, other: object) -> bool:
        return (
            isinstance(other, ExternalConnectionNode)
            and self.conn_type == other.conn_type
            and self.node_id   == other.node_id
        )

    def __hash__(self) -> int:
        return hash((self.conn_type, self.node_id))

    def __repr__(self) -> str:
        return (
            f"ExternalConnectionNode({self.conn_type!r}, "
            f"{self.name!r}, {self.source!r})"
        )


# ── Private XML helpers ───────────────────────────────────────────────────

def _read_zip_xml(zf: zipfile.ZipFile, path: str) -> ET.Element | None:
    """Return parsed XML root from a zip member, or None if not present/empty."""
    try:
        content = zf.read(path)
        if not content.strip():
            return None
        return ET.fromstring(content)
    except KeyError:
        return None
    except ET.ParseError:
        return None


def _parse_xml_rels(zf: zipfile.ZipFile, rels_path: str) -> dict[str, dict]:
    """Return {rId: {target, type}} from an Office .rels file."""
    root = _read_zip_xml(zf, rels_path)
    if root is None:
        return {}
    return {
        rel.get("Id", ""): {
            "target": rel.get("Target", ""),
            "type":   rel.get("Type",   ""),
        }
        for rel in root
    }


def _extract_source_from_connstr(conn_str: str) -> str:
    """Pull a meaningful label out of an ODBC/OLE DB connection string."""
    for pattern in (
        r"DSN=([^;]+)",
        r"Data Source=([^;$]+)",
        r"Server=([^;]+)",
        r"Host=([^;]+)",
    ):
        m = re.search(pattern, conn_str, re.I)
        if m:
            return m.group(1).strip()
    # Power Query Location= points to the query name, not a server
    m = re.search(r"Location=([^;]+)", conn_str, re.I)
    if m:
        loc = m.group(1).strip()
        if loc != "$Workbook$":
            return loc
    return ""


def _parse_connections_xml(
    zf: zipfile.ZipFile,
) -> dict[str, "ExternalConnectionNode"]:
    """
    Parse ``xl/connections.xml`` and return ``{conn_id: ExternalConnectionNode}``.

    Detects:
    - ODBC  (type=1)
    - OLE DB (type=5)
    - Power Query (OLE DB + Mashup provider OR x15 power-query ext)
    - Web queries (type=4)
    """
    root = _read_zip_xml(zf, "xl/connections.xml")
    if root is None:
        return {}

    ns = _OOXML_NS
    result: dict[str, ExternalConnectionNode] = {}

    for conn in root.iter(f"{{{ns}}}connection"):
        conn_id   = conn.get("id", "")
        name      = conn.get("name", conn_id)
        type_int  = int(conn.get("type", "0"))

        # ── Detect Power Query via extLst URI ─────────────────────────────
        is_pq = False
        ext_lst = conn.find(f"{{{ns}}}extLst")
        if ext_lst is not None:
            for ext in ext_lst:
                if _PQ_EXT_URI in ext.get("uri", ""):
                    is_pq = True
                    break

        # ── Get connection string / URL ────────────────────────────────────
        source = ""
        db_pr  = conn.find(f"{{{ns}}}dbPr")
        if db_pr is not None:
            conn_str = db_pr.get("connection", "")
            if "Mashup" in conn_str:
                is_pq = True
            source = _extract_source_from_connstr(conn_str)

        web_pr = conn.find(f"{{{ns}}}webPr")
        if web_pr is not None:
            source = web_pr.get("url", name)

        # ── Classify ──────────────────────────────────────────────────────
        if is_pq:
            # Strip Excel's "Query - " prefix that auto-generated names get
            clean_name = re.sub(r"^Query\s*-\s*", "", name, flags=re.I).strip()
            result[conn_id] = ExternalConnectionNode(
                "powerquery", clean_name, source or clean_name
            )
        elif type_int == _TYPE_WEB or web_pr is not None:
            result[conn_id] = ExternalConnectionNode("web", name, source or name)
        elif type_int in (_TYPE_ODBC, _TYPE_DAO):
            result[conn_id] = ExternalConnectionNode("odbc", name, source or name)
        elif type_int == _TYPE_OLEDB:
            result[conn_id] = ExternalConnectionNode("oledb", name, source or name)
        # Skip type 0 / unknown

    return result


def _get_sheet_xml_paths(zf: zipfile.ZipFile) -> dict[str, str]:
    """
    Return ``{sheet_name: 'xl/worksheets/sheetN.xml'}`` by reading
    workbook.xml and its .rels file.
    """
    wb_root = _read_zip_xml(zf, "xl/workbook.xml")
    if wb_root is None:
        return {}

    rels = _parse_xml_rels(zf, "xl/_rels/workbook.xml.rels")
    ns   = _OOXML_NS

    sheet_paths: dict[str, str] = {}
    for sheet_el in wb_root.iter(f"{{{ns}}}sheet"):
        sname = sheet_el.get("name", "")
        rid   = sheet_el.get(
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id",
            "",
        )
        rel   = rels.get(rid, {})
        target = rel.get("target", "")
        if target:
            # Target is relative to xl/
            full = f"xl/{target}" if not target.startswith("xl/") else target
            sheet_paths[sname] = full
    return sheet_paths


def _conn_ids_for_sheet(
    zf: zipfile.ZipFile,
    sheet_xml_path: str,
) -> list[str]:
    """
    Return all connection IDs used by a sheet, scanning:
    - queryTable XML files (linked via queryTablePart relationships)
    - pivot-cache definition XML files (linked via pivotTable relationships)
    """
    ns    = _OOXML_NS
    qt_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    sheet_root = _read_zip_xml(zf, sheet_xml_path)
    if sheet_root is None:
        return []

    # Build the sheet's relationship map
    sheet_file = sheet_xml_path.rsplit("/", 1)[-1]   # e.g. sheet1.xml
    rels_path  = f"xl/worksheets/_rels/{sheet_file}.rels"
    rels       = _parse_xml_rels(zf, rels_path)

    conn_ids: list[str] = []

    # ── Query tables ──────────────────────────────────────────────────────
    for qt_part in sheet_root.iter(f"{{{ns}}}queryTablePart"):
        rid = qt_part.get(
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id",
            "",
        )
        rel = rels.get(rid, {})
        target = rel.get("target", "")
        if not target:
            continue
        # target is like "../queryTables/queryTable1.xml"
        qt_path = "xl/" + target.lstrip("../")
        qt_root = _read_zip_xml(zf, qt_path)
        if qt_root is not None:
            cid = qt_root.get("connectionId", "")
            if cid:
                conn_ids.append(cid)

    # ── Pivot tables → pivot-cache definitions ────────────────────────────
    for pt_part in sheet_root.iter(f"{{{ns}}}pivotTablePart"):
        rid = pt_part.get(
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id",
            "",
        )
        rel = rels.get(rid, {})
        target = rel.get("target", "")
        if not target:
            continue
        pt_path  = "xl/" + target.lstrip("../")
        pt_root  = _read_zip_xml(zf, pt_path)
        if pt_root is None:
            continue
        cache_id = pt_root.get("cacheId", "")
        if not cache_id:
            continue
        # Resolve pivot cache definition via workbook rels
        wb_rels = _parse_xml_rels(zf, "xl/_rels/workbook.xml.rels")
        cache_type = (
            "http://schemas.openxmlformats.org/officeDocument/2006/"
            "relationships/pivotCacheDefinition"
        )
        for rel_info in wb_rels.values():
            if cache_type not in rel_info.get("type", ""):
                continue
            cache_path = "xl/" + rel_info["target"].lstrip("../")
            cache_root = _read_zip_xml(zf, cache_path)
            if cache_root is None:
                continue
            # Pivot cache def has <cacheSource type="worksheet" connectionId="N"/>
            for cs in cache_root.iter(f"{{{_OOXML_NS}}}cacheSource"):
                cid = cs.get("connectionId", "")
                if cid:
                    conn_ids.append(cid)

    return conn_ids


# ── Public extraction function ────────────────────────────────────────────

def extract_external_connections(
    path: str,
    sheet_nodes: set[SheetNode],
) -> tuple[list["ExternalConnectionNode"], list[tuple["ExternalConnectionNode", SheetNode]]]:
    """
    Scan the workbook zip for ODBC, Power Query, and web-query connections.

    Returns
    -------
    conn_nodes : list[ExternalConnectionNode]
        Every connection found — shown in the diagram even without edges.
    conn_edges : list[(ExternalConnectionNode, SheetNode)]
        Only connections that can be traced to a specific sheet via
        queryTable or pivotTable XML.  Unlinked connections are included
        in *conn_nodes* but have no edge — they appear as floating upstream
        nodes without arrows, avoiding spurious "connected to everything" noise.
    """
    conn_nodes: list[ExternalConnectionNode] = []
    conn_edges: list[tuple[ExternalConnectionNode, SheetNode]] = []

    try:
        zf_handle = zipfile.ZipFile(path, "r")
    except (zipfile.BadZipFile, OSError):
        return [], []

    with zf_handle as zf:
        conn_map = _parse_connections_xml(zf)
        if not conn_map:
            return [], []

        cur_nodes = [n for n in sheet_nodes if n.workbook is None]
        sheet_by_name = {n.sheet: n for n in cur_nodes}

        # Map sheet name → xl/worksheets/sheetN.xml
        sheet_xml_paths = _get_sheet_xml_paths(zf)

        # conn_id → list of sheets that use it
        conn_to_sheets: dict[str, list[SheetNode]] = {}
        for sheet_name, xml_path in sheet_xml_paths.items():
            node = sheet_by_name.get(sheet_name)
            if node is None:
                continue
            for cid in _conn_ids_for_sheet(zf, xml_path):
                conn_to_sheets.setdefault(cid, []).append(node)

    seen_nodes: set[str] = set()
    seen_edges: set[tuple[str, str]] = set()

    def _add_node(conn_node: ExternalConnectionNode) -> None:
        if conn_node.node_id not in seen_nodes:
            seen_nodes.add(conn_node.node_id)
            conn_nodes.append(conn_node)

    def _add_edge(conn_node: ExternalConnectionNode, sheet_node: SheetNode) -> None:
        _add_node(conn_node)
        key = (conn_node.node_id, sheet_node.node_id)
        if key not in seen_edges:
            seen_edges.add(key)
            conn_edges.append((conn_node, sheet_node))

    for cid, conn_node in conn_map.items():
        targets = conn_to_sheets.get(cid)
        if targets:
            for sn in targets:
                _add_edge(conn_node, sn)
        else:
            # No queryTable/pivotTable links this connection to a specific sheet.
            # Show it in the diagram as a floating upstream node (no arrows).
            _add_node(conn_node)

    return conn_nodes, conn_edges


def extract_upstream_file_connections(
    main_path: str,
    all_nodes: set[SheetNode],
) -> dict[str, list["ExternalConnectionNode"]]:
    """
    For every external workbook referenced in *all_nodes* (i.e. nodes whose
    ``workbook`` attribute is not None), check whether that file exists in
    the **same directory** as *main_path*.  If it does, extract its ODBC /
    Power Query / Web connections — but **not** its formula-level sheet
    dependencies.

    Returns ``{workbook_filename: [ExternalConnectionNode]}`` only for files
    that are both present on disk and have at least one detectable connection.
    Connection node IDs are scoped to the workbook (via ``wb_scope``) so
    identically-named connections in different files never collide.

    Only one level of depth is traversed — upstream files of upstream files
    are not examined.
    """
    main_dir = os.path.dirname(os.path.abspath(main_path))

    # Unique external workbook names from formula references
    upstream_wb_names: set[str] = {
        n.workbook for n in all_nodes if n.workbook is not None
    }
    if not upstream_wb_names:
        return {}

    result: dict[str, list[ExternalConnectionNode]] = {}

    for wb_name in sorted(upstream_wb_names):  # sorted for reproducibility
        candidate = os.path.join(main_dir, wb_name)
        if not os.path.isfile(candidate):
            continue  # file not in same directory — skip silently

        try:
            with zipfile.ZipFile(candidate, "r") as zf:
                raw_conns = _parse_connections_xml(zf)
        except Exception:
            continue  # corrupt / not a real xlsx — skip

        if not raw_conns:
            continue

        scoped: list[ExternalConnectionNode] = []
        seen_ids: set[str] = set()
        for cn in raw_conns.values():
            scoped_cn = ExternalConnectionNode(
                cn.conn_type, cn.name, cn.source, wb_scope=wb_name
            )
            if scoped_cn.node_id not in seen_ids:
                seen_ids.add(scoped_cn.node_id)
                scoped.append(scoped_cn)

        if scoped:
            result[wb_name] = scoped

    return result


# ---------------------------------------------------------------------------
# Dependency-graph builder
# ---------------------------------------------------------------------------

def build_dependency_graph(
    workbook_path: str,
) -> tuple[set[SheetNode], list[tuple[SheetNode, SheetNode]]]:
    """
    Open *workbook_path* and trace formula-level sheet dependencies.

    Parameters
    ----------
    workbook_path : str
        Path to a ``.xlsx`` / ``.xlsm`` file.

    Returns
    -------
    nodes : set[SheetNode]
        All sheets discovered (current-workbook + any upstream sheets found
        in formulas).
    edges : list[(src, dst)]
        One entry per *unique* (source-sheet, destination-sheet) pair.
        ``src → dst`` means "dst reads from src".
        Multiple formulas between the same pair produce exactly one edge.
    """
    wb_path = str(workbook_path)
    current_filename = os.path.basename(wb_path)

    # data_only=False → read formula strings, not cached values.
    # read_only=False (default) → returns Worksheet objects with full API;
    # read_only=True returns ReadOnlyWorksheet which has a different type,
    # so we keep the default to avoid isinstance confusion.
    wb = load_workbook(wb_path, data_only=False)
    try:
        sheet_names: set[str] = set(wb.sheetnames)
        nodes: set[SheetNode] = set()
        edge_set: set[tuple[SheetNode, SheetNode]] = set()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Add every worksheet in the current file as a node even if it
            # has no formulas (isolated nodes appear in the diagram).
            # Use duck-typing: chart/macro sheets lack iter_rows.
            if not hasattr(ws, "iter_rows"):
                nodes.add(SheetNode(workbook=None, sheet=sheet_name))
                continue

            dst_node = SheetNode(workbook=None, sheet=sheet_name)
            nodes.add(dst_node)

            for row in ws.iter_rows():
                for cell in row:
                    formula = cell.value
                    if not isinstance(formula, str) or not formula.startswith("="):
                        continue

                    for wb_ref, sh_ref in extract_refs(formula):
                        # ── Determine source node ──────────────────────────
                        if wb_ref is not None:
                            # Treat a reference to *this same file* by name
                            # as an internal reference.
                            if wb_ref.lower() == current_filename.lower():
                                if sh_ref not in sheet_names:
                                    continue
                                src_node = SheetNode(workbook=None, sheet=sh_ref)
                            else:
                                src_node = SheetNode(
                                    workbook=wb_ref, sheet=sh_ref
                                )
                        else:
                            # Internal reference
                            if sh_ref not in sheet_names:
                                continue  # reference to an unknown sheet name
                            if sh_ref == sheet_name:
                                continue  # skip self-references
                            src_node = SheetNode(workbook=None, sheet=sh_ref)

                        nodes.add(src_node)
                        edge_set.add((src_node, dst_node))
    finally:
        wb.close()

    return nodes, list(edge_set)


# ---------------------------------------------------------------------------
# Mermaid diagram builder
# ---------------------------------------------------------------------------

def build_mermaid(
    nodes: set[SheetNode],
    edges: list[tuple[SheetNode, SheetNode]],
    current_file: str,
    conn_nodes: list["ExternalConnectionNode"] | None = None,
    conn_edges: list[tuple["ExternalConnectionNode", SheetNode]] | None = None,
    upstream_conn_map: dict[str, list["ExternalConnectionNode"]] | None = None,
) -> str:
    """
    Produce a ``flowchart LR`` Mermaid diagram string.

    Sheets from the same workbook are wrapped in a labelled subgraph.
    The current file's sheets appear first; external workbooks follow in
    alphabetical order.

    *conn_nodes* — current file's ExternalConnectionNodes (including those
    with no arrows). *conn_edges* — only the current-file ones with a
    confirmed queryTable/pivotTable target sheet.

    *upstream_conn_map* — ``{workbook_filename: [ExternalConnectionNode]}``
    for upstream Excel files found on disk.  Their connection nodes are
    rendered **inside** that workbook's subgraph (no arrows) so it's clear
    which file they belong to, without cluttering the diagram.

    Node colours:
    - ODBC   → warm orange  (``#f4a261``)
    - OLE DB → amber        (``#e9c46a``)
    - Power Query → forest green (``#52b788``)
    - Web    → violet       (``#c77dff``)
    """
    lines = ["flowchart LR"]

    # ── Collect current-file connection nodes ─────────────────────────────
    # Start with explicitly provided nodes; supplement from conn_edges so
    # callers that only pass conn_edges still get subgraphs.
    cur_conn_nodes = list(conn_nodes or [])
    if conn_edges:
        seen_ids = {cn.node_id for cn in cur_conn_nodes}
        for cn, _ in conn_edges:
            if cn.node_id not in seen_ids:
                seen_ids.add(cn.node_id)
                cur_conn_nodes.append(cn)

    # All connection nodes (current file + upstream files) — used for classDef
    all_conn_nodes: list[ExternalConnectionNode] = list(cur_conn_nodes)
    if upstream_conn_map:
        for cn_list in upstream_conn_map.values():
            all_conn_nodes.extend(cn_list)

    # ── classDef for connection node colours ─────────────────────────────
    conn_type_order = ["odbc", "oledb", "powerquery", "web"]
    conn_type_labels = {
        "odbc":       "ODBC",
        "oledb":      "OLE DB",
        "powerquery": "Power Query",
        "web":        "Web",
    }

    if all_conn_nodes:
        lines += [
            "  classDef odbcNode fill:#f4a261,stroke:#c1440e,color:#1a1a2e",
            "  classDef oledbNode fill:#e9c46a,stroke:#b5830d,color:#1a1a2e",
            "  classDef pqNode fill:#52b788,stroke:#1b4332,color:#fff",
            "  classDef webNode fill:#c77dff,stroke:#7b2d8b,color:#fff",
        ]

    # ── Current-file connection subgraphs (before sheet subgraphs) ────────
    cur_by_type: dict[str, list[ExternalConnectionNode]] = {}
    for cn in cur_conn_nodes:
        cur_by_type.setdefault(cn.conn_type, []).append(cn)

    for ct in conn_type_order:
        if ct not in cur_by_type:
            continue
        sg_label = conn_type_labels.get(ct, ct)
        sg_id    = f"SG_CONN_{ct.upper()}"
        lines.append(f'  subgraph {sg_id}["{sg_label} Connections"]')
        for cn in sorted(cur_by_type[ct], key=lambda c: c.node_id):
            safe_label = cn.label.replace('"', "'")
            lines.append(f'    {cn.node_id}["{safe_label}"]')
        lines.append("  end")

    # ── Group sheet nodes by workbook ─────────────────────────────────────
    by_wb: dict[str | None, list[SheetNode]] = {}
    for node in nodes:
        by_wb.setdefault(node.workbook, []).append(node)

    def _wb_sort_key(wb: str | None) -> str:
        return "" if wb is None else wb.lower()

    # ── Emit sheet subgraphs ──────────────────────────────────────────────
    for wb_name in sorted(by_wb.keys(), key=_wb_sort_key):
        sheets = sorted(by_wb[wb_name], key=lambda n: n.sheet.lower())
        display_name = current_file if wb_name is None else wb_name

        if wb_name is None:
            sg_id = "SG_CUR"
        else:
            sg_id = "SG_EXT__" + re.sub(r"[^A-Za-z0-9_]", "_", wb_name)

        safe_label = display_name.replace('"', "'")
        lines.append(f'  subgraph {sg_id}["{safe_label}"]')
        for node in sheets:
            safe_sheet = node.sheet.replace('"', "'")
            lines.append(f'    {node.node_id}["{safe_sheet}"]')

        # Upstream-file connections appear inside this workbook's subgraph.
        # No arrows are drawn — the subgraph context makes the relationship
        # clear without crowding the diagram.
        if wb_name is not None and upstream_conn_map and wb_name in upstream_conn_map:
            for cn in sorted(upstream_conn_map[wb_name], key=lambda c: c.node_id):
                safe_cn_label = cn.label.replace('"', "'")
                lines.append(f'    {cn.node_id}["{safe_cn_label}"]')

        lines.append("  end")

    # ── Emit sheet-to-sheet edges ─────────────────────────────────────────
    if edges:
        lines.append("")
    for src, dst in sorted(edges, key=lambda e: (e[0].node_id, e[1].node_id)):
        lines.append(f"  {src.node_id} --> {dst.node_id}")

    # ── Emit connection → sheet edges (current file only) ─────────────────
    if conn_edges:
        lines.append("")
        for cn, sn in sorted(conn_edges, key=lambda e: (e[0].node_id, e[1].node_id)):
            lines.append(f"  {cn.node_id} --> {sn.node_id}")

    # ── Apply CSS classes to ALL connection nodes ─────────────────────────
    if all_conn_nodes:
        all_by_type: dict[str, list[ExternalConnectionNode]] = {}
        for cn in all_conn_nodes:
            all_by_type.setdefault(cn.conn_type, []).append(cn)
        lines.append("")
        for ct in conn_type_order:
            if ct not in all_by_type:
                continue
            cls = all_by_type[ct][0].mermaid_class
            ids = ",".join(
                cn.node_id
                for cn in sorted(all_by_type[ct], key=lambda c: c.node_id)
            )
            lines.append(f"  class {ids} {cls}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# HTML generation
# ---------------------------------------------------------------------------

def generate_html(mermaid_text: str, title: str = "Excel Lineage") -> str:
    """Return a complete, self-contained interactive HTML page.

    Delegates to mermaid_interactive.generate_html() so the full
    pan/zoom, click-to-highlight, and PNG-download engine is always in
    sync with the parent module.  Only the <title> tag is overridden to
    show the workbook filename.
    """
    html = _mi_generate_html(mermaid_text)
    safe_title = html_module.escape(title)
    return html.replace(
        "<title>Mermaid Diagram</title>",
        f"<title>Lineage \u2014 {safe_title}</title>",
    )

# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

_SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        prog="excel_lineage",
        description=(
            "Generate a sheet-level dependency diagram from an Excel workbook."
        ),
    )
    p.add_argument("input", help="Path to .xlsx / .xlsm workbook")
    p.add_argument(
        "-o", "--output",
        default=None,
        help="Output HTML path (default: <workbook>_lineage.html)",
    )
    args = p.parse_args(argv)

    wb_path = args.input
    if not os.path.isfile(wb_path):
        print(f"Error: file not found: {wb_path!r}", file=sys.stderr)
        return 1

    ext = Path(wb_path).suffix.lower()
    if ext not in _SUPPORTED_EXTENSIONS:
        print(
            f"Error: unsupported format {ext!r}. "
            f"Supported: {', '.join(sorted(_SUPPORTED_EXTENSIONS))}",
            file=sys.stderr,
        )
        return 1

    output_path = args.output or (
        str(Path(wb_path).with_suffix("")) + "_lineage.html"
    )

    print(f"Reading {wb_path!r} ...", file=sys.stderr)
    try:
        nodes, edges = build_dependency_graph(wb_path)
    except Exception as exc:
        print(f"Error reading workbook: {exc}", file=sys.stderr)
        return 1

    title = os.path.basename(wb_path)

    if not edges:
        print(
            "Warning: no cross-sheet formula dependencies found. "
            "Diagram will show sheets without connections.",
            file=sys.stderr,
        )

    conn_nodes, conn_edges = extract_external_connections(wb_path, nodes)
    if conn_nodes:
        linked = len({cn for cn, _ in conn_edges})
        total  = len(conn_nodes)
        print(
            f"Found {total} external connection(s) (ODBC / Power Query / Web); "
            f"{linked} linked to specific sheet(s).",
            file=sys.stderr,
        )

    upstream_conn_map = extract_upstream_file_connections(wb_path, nodes)
    if upstream_conn_map:
        n_files = len(upstream_conn_map)
        n_conns = sum(len(v) for v in upstream_conn_map.values())
        print(
            f"Found {n_files} upstream file(s) on disk with "
            f"{n_conns} external connection(s) total.",
            file=sys.stderr,
        )

    mermaid_text  = build_mermaid(
        nodes, edges, current_file=title,
        conn_nodes=conn_nodes, conn_edges=conn_edges,
        upstream_conn_map=upstream_conn_map,
    )
    html_content  = generate_html(mermaid_text, title=title)

    try:
        with open(output_path, "w", encoding="utf-8") as fh:
            fh.write(html_content)
    except OSError as exc:
        print(f"Error writing output: {exc}", file=sys.stderr)
        return 1

    print(f"Lineage diagram written to: {os.path.abspath(output_path)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
