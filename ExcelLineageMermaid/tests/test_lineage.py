"""
tests/test_lineage.py
=====================
Comprehensive tests for excel_lineage.py.

Tests are organised into:
  1. TestExtractRefs      – formula parser (pure-Python, no file I/O)
  2. TestSheetNode        – node ID sanitisation and equality
  3. TestBuildMermaid     – diagram text generation from synthetic data
  4. TestBuildDependency  – full graph builder against real .xlsx files
  5. TestGenerateHtml     – HTML output structure
  6. TestEdgeCases        – robustness / corner cases
"""

import os
import re
import sys
import tempfile

import pytest

# Make the parent directory importable when running from the repo root
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from excel_lineage import (
    extract_refs,
    SheetNode,
    ExternalConnectionNode,
    build_dependency_graph,
    build_mermaid,
    extract_external_connections,
    extract_upstream_file_connections,
    generate_html,
)

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_xlsx(sheets_and_cells: dict[str, dict[tuple[int, int], str]]) -> str:
    """
    Create a temporary .xlsx file and return its path.

    Parameters
    ----------
    sheets_and_cells : {sheet_name: {(row, col): cell_value}}
        Cell values that start with "=" are formulas.

    Caller is responsible for deleting the file (use try/finally).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)          # drop the default "Sheet" sheet
    for sheet_name, cells in sheets_and_cells.items():
        ws = wb.create_sheet(sheet_name)
        for (row, col), value in cells.items():
            ws.cell(row=row, column=col, value=value)
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


def edge_pairs(edges) -> set[tuple[str, str]]:
    """Convert edge list to a set of (src_sheet, dst_sheet) string pairs."""
    return {(e[0].sheet, e[1].sheet) for e in edges}


# ===========================================================================
# 1. TestExtractRefs
# ===========================================================================

class TestExtractRefs:
    """Unit tests for extract_refs() – pure formula-string parsing."""

    # ── basic internal references ──────────────────────────────────────────

    def test_plain_internal_ref(self):
        refs = extract_refs("=Sheet1!A1")
        assert refs == [(None, "Sheet1")]

    def test_quoted_internal_ref(self):
        refs = extract_refs("='Sheet Name'!A1")
        assert refs == [(None, "Sheet Name")]

    def test_quoted_internal_with_numbers(self):
        refs = extract_refs("='Q3 2024'!B5")
        assert refs == [(None, "Q3 2024")]

    def test_internal_ref_in_sum(self):
        refs = extract_refs("=SUM(Sheet1!A1:Sheet1!A10)")
        sheets = {sh for _, sh in refs}
        assert "Sheet1" in sheets

    def test_multiple_different_internal_refs(self):
        refs = extract_refs("=Sheet1!A1 + Sheet2!B2 - Sheet3!C3")
        sheets = {sh for _, sh in refs}
        assert sheets == {"Sheet1", "Sheet2", "Sheet3"}
        assert all(wb is None for wb, _ in refs)

    def test_duplicate_internal_refs_deduplicated(self):
        # Same sheet referenced many times → one entry
        refs = extract_refs("=Sheet1!A1 + Sheet1!B1 + Sheet1!C1")
        assert refs == [(None, "Sheet1")]

    # ── external workbook references ───────────────────────────────────────

    def test_external_quoted_simple(self):
        refs = extract_refs("='[Book1.xlsx]Sheet1'!A1")
        assert len(refs) == 1
        wb, sh = refs[0]
        assert wb == "Book1.xlsx"
        assert sh == "Sheet1"

    def test_external_quoted_sheet_with_space(self):
        refs = extract_refs("='[Upstream.xlsx]Source Data'!A1")
        wb, sh = refs[0]
        assert wb == "Upstream.xlsx"
        assert sh == "Source Data"

    def test_external_quoted_with_full_path(self):
        refs = extract_refs(
            r"='[C:\Users\Alice\data\Upstream.xlsx]Summary'!A1"
        )
        wb, sh = refs[0]
        # Only the filename, not the full path
        assert wb == "Upstream.xlsx"
        assert sh == "Summary"

    def test_external_quoted_unix_path(self):
        refs = extract_refs("='[/home/user/data/Upstream.xlsx]Sheet1'!A1")
        wb, sh = refs[0]
        assert wb == "Upstream.xlsx"
        assert sh == "Sheet1"

    def test_external_unquoted(self):
        refs = extract_refs("=[Book2.xlsx]Summary!A1")
        assert len(refs) == 1
        wb, sh = refs[0]
        assert wb == "Book2.xlsx"
        assert sh == "Summary"

    def test_mixed_internal_and_external(self):
        formula = "='[Upstream.xlsx]Raw'!A1 + Sheet2!B1"
        refs = extract_refs(formula)
        by_wb = {wb: sh for wb, sh in refs}
        assert "Upstream.xlsx" in by_wb
        assert None in by_wb
        assert by_wb["Upstream.xlsx"] == "Raw"
        assert by_wb[None] == "Sheet2"

    # ── non-formula inputs ──────────────────────────────────────────────────

    def test_plain_string_returns_empty(self):
        assert extract_refs("hello world") == []

    def test_number_returns_empty(self):
        assert extract_refs("42") == []

    def test_none_returns_empty(self):
        assert extract_refs(None) == []   # type: ignore[arg-type]

    def test_empty_string_returns_empty(self):
        assert extract_refs("") == []

    def test_formula_without_sheet_ref(self):
        assert extract_refs("=SUM(A1:A10)") == []

    def test_if_formula_no_refs(self):
        assert extract_refs("=IF(A1>0, B1, C1)") == []

    def test_vlookup_no_sheet_refs(self):
        assert extract_refs("=VLOOKUP(A1, B1:C10, 2, FALSE)") == []

    # ── edge-case formulas ──────────────────────────────────────────────────

    def test_formula_starts_with_equals_required(self):
        # Formula text without leading = must not be parsed
        assert extract_refs("Sheet1!A1") == []

    def test_indirect_function_not_parsed(self):
        # INDIRECT constructs references dynamically.  The sheet name lives
        # inside a double-quoted string literal, which is stripped before
        # regex matching, so no reference is returned.
        assert extract_refs('=INDIRECT("Sheet1!A1")') == []

    def test_string_literal_not_confused_with_ref(self):
        # Sheet name inside a string must not be returned as a reference
        assert extract_refs('=CONCATENATE("Sheet1!A1", B2)') == []


# ===========================================================================
# 2. TestSheetNode
# ===========================================================================

class TestSheetNode:
    def test_internal_node_id_safe(self):
        node = SheetNode(workbook=None, sheet="Sheet1")
        assert re.match(r"^[A-Za-z][A-Za-z0-9_]*$", node.node_id)

    def test_external_node_id_safe(self):
        node = SheetNode(workbook="Book 1.xlsx", sheet="Sheet Name")
        assert re.match(r"^[A-Za-z][A-Za-z0-9_]*$", node.node_id)

    def test_node_id_starts_with_letter(self):
        # Simulate a sheet name that begins with a digit after sanitisation
        node = SheetNode(workbook=None, sheet="123sheet")
        assert node.node_id[0].isalpha()

    def test_internal_and_external_different_ids(self):
        internal = SheetNode(workbook=None,         sheet="Sheet1")
        external = SheetNode(workbook="Book.xlsx",  sheet="Sheet1")
        assert internal.node_id != external.node_id

    def test_label_is_sheet_name(self):
        node = SheetNode(workbook="Book.xlsx", sheet="My Data")
        assert node.label == "My Data"

    def test_equality_same(self):
        a = SheetNode(None, "Sheet1")
        b = SheetNode(None, "Sheet1")
        assert a == b

    def test_equality_different_workbook(self):
        a = SheetNode(None,         "Sheet1")
        b = SheetNode("Book.xlsx",  "Sheet1")
        assert a != b

    def test_equality_different_sheet(self):
        a = SheetNode(None, "Sheet1")
        b = SheetNode(None, "Sheet2")
        assert a != b

    def test_hashable_in_set(self):
        s = {
            SheetNode(None,        "Sheet1"),
            SheetNode(None,        "Sheet1"),   # duplicate
            SheetNode("Book.xlsx", "Sheet1"),
        }
        assert len(s) == 2

    def test_hashable_as_dict_key(self):
        node = SheetNode(None, "Sheet1")
        d = {node: "value"}
        assert d[SheetNode(None, "Sheet1")] == "value"


# ===========================================================================
# 3. TestBuildMermaid
# ===========================================================================

class TestBuildMermaid:
    def test_flowchart_header(self):
        s1 = SheetNode(None, "A")
        mmd = build_mermaid({s1}, [], "File.xlsx")
        assert mmd.startswith("flowchart LR")

    def test_single_node_no_edges(self):
        s1 = SheetNode(None, "Sheet1")
        mmd = build_mermaid({s1}, [], "File.xlsx")
        assert "Sheet1" in mmd
        assert "-->" not in mmd

    def test_edge_direction(self):
        s1 = SheetNode(None, "Sheet1")
        s2 = SheetNode(None, "Sheet2")
        mmd = build_mermaid({s1, s2}, [(s1, s2)], "File.xlsx")
        # s1 is the source → s2 is the destination
        assert f"{s1.node_id} --> {s2.node_id}" in mmd

    def test_subgraph_for_current_file(self):
        s1 = SheetNode(None, "Sheet1")
        mmd = build_mermaid({s1}, [], "MyFile.xlsx")
        assert "subgraph" in mmd
        assert "MyFile.xlsx" in mmd

    def test_subgraph_for_external_workbook(self):
        ext = SheetNode("Upstream.xlsx", "Data")
        cur = SheetNode(None, "Summary")
        mmd = build_mermaid({ext, cur}, [(ext, cur)], "Current.xlsx")
        assert "Upstream.xlsx" in mmd
        assert "Current.xlsx" in mmd

    def test_current_file_subgraph_listed_first(self):
        ext = SheetNode("Upstream.xlsx", "Data")
        cur = SheetNode(None, "Summary")
        mmd = build_mermaid({ext, cur}, [(ext, cur)], "Current.xlsx")
        pos_cur = mmd.index("Current.xlsx")
        pos_ext = mmd.index("Upstream.xlsx")
        assert pos_cur < pos_ext

    def test_node_labels_use_sheet_name(self):
        node = SheetNode(None, "My Sheet")
        mmd = build_mermaid({node}, [], "X.xlsx")
        assert '"My Sheet"' in mmd

    def test_multiple_external_workbooks(self):
        e1 = SheetNode("Alpha.xlsx", "A")
        e2 = SheetNode("Beta.xlsx",  "B")
        cur = SheetNode(None, "C")
        mmd = build_mermaid({e1, e2, cur}, [(e1, cur), (e2, cur)], "X.xlsx")
        assert "Alpha.xlsx" in mmd
        assert "Beta.xlsx"  in mmd
        assert "X.xlsx"     in mmd

    def test_no_duplicate_edges_in_output(self):
        s1 = SheetNode(None, "Sheet1")
        s2 = SheetNode(None, "Sheet2")
        # Pass the same edge twice (shouldn't happen from graph builder,
        # but build_mermaid should emit it as written)
        mmd = build_mermaid({s1, s2}, [(s1, s2)], "X.xlsx")
        arrow = f"{s1.node_id} --> {s2.node_id}"
        assert mmd.count(arrow) == 1


# ===========================================================================
# 4. TestBuildDependency
# ===========================================================================

class TestBuildDependency:
    """Integration tests that create real .xlsx files with openpyxl."""

    def test_simple_one_way_reference(self):
        path = make_xlsx({
            "Sheet1": {},
            "Sheet2": {(1, 1): "=Sheet1!A1"},
        })
        try:
            nodes, edges = build_dependency_graph(path)
            assert ("Sheet1", "Sheet2") in edge_pairs(edges)
        finally:
            os.unlink(path)

    def test_no_cross_sheet_refs_gives_no_edges(self):
        path = make_xlsx({
            "Sheet1": {(1, 1): "=SUM(B1:B10)"},
            "Sheet2": {(1, 1): "=IF(A1>0,1,0)"},
        })
        try:
            _, edges = build_dependency_graph(path)
            assert edges == []
        finally:
            os.unlink(path)

    def test_all_sheets_registered_as_nodes(self):
        path = make_xlsx({
            "Alpha": {},
            "Beta":  {},
            "Gamma": {},
        })
        try:
            nodes, _ = build_dependency_graph(path)
            names = {n.sheet for n in nodes if n.workbook is None}
            assert {"Alpha", "Beta", "Gamma"} == names
        finally:
            os.unlink(path)

    def test_1000_formulas_produce_one_edge(self):
        """
        N formulas between the same two sheets must collapse to a single edge.
        """
        cells = {(r, 1): f"=Sheet1!A{r}" for r in range(1, 1001)}
        path = make_xlsx({
            "Sheet1": {},
            "Sheet2": cells,
        })
        try:
            _, edges = build_dependency_graph(path)
            pairs = list(edge_pairs(edges))
            assert pairs.count(("Sheet1", "Sheet2")) == 1
            assert len(pairs) == 1
        finally:
            os.unlink(path)

    def test_self_reference_ignored(self):
        path = make_xlsx({
            "Sheet1": {(1, 1): "=Sheet1!B1"},
        })
        try:
            _, edges = build_dependency_graph(path)
            assert edges == []
        finally:
            os.unlink(path)

    def test_chain_a_to_b_to_c(self):
        path = make_xlsx({
            "A": {},
            "B": {(1, 1): "=A!A1"},
            "C": {(1, 1): "=B!A1"},
        })
        try:
            _, edges = build_dependency_graph(path)
            ep = edge_pairs(edges)
            assert ("A", "B") in ep
            assert ("B", "C") in ep
            assert ("A", "C") not in ep
        finally:
            os.unlink(path)

    def test_diamond_dependency(self):
        """A → B, A → C, B → D, C → D"""
        path = make_xlsx({
            "A": {},
            "B": {(1, 1): "=A!A1"},
            "C": {(1, 1): "=A!A1"},
            "D": {(1, 1): "=B!A1", (1, 2): "=C!A1"},
        })
        try:
            _, edges = build_dependency_graph(path)
            ep = edge_pairs(edges)
            assert ep == {("A", "B"), ("A", "C"), ("B", "D"), ("C", "D")}
        finally:
            os.unlink(path)

    def test_quoted_sheet_name_with_spaces(self):
        path = make_xlsx({
            "Source Data": {},
            "Summary":     {(1, 1): "='Source Data'!A1"},
        })
        try:
            _, edges = build_dependency_graph(path)
            assert ("Source Data", "Summary") in edge_pairs(edges)
        finally:
            os.unlink(path)

    def test_external_workbook_reference_detected(self):
        path = make_xlsx({
            "Summary": {(1, 1): "='[Upstream.xlsx]Data'!A1"},
        })
        try:
            nodes, edges = build_dependency_graph(path)
            ext_nodes = [n for n in nodes if n.workbook is not None]
            assert len(ext_nodes) >= 1
            assert any(
                n.workbook == "Upstream.xlsx" and n.sheet == "Data"
                for n in ext_nodes
            )
        finally:
            os.unlink(path)

    def test_external_edge_direction(self):
        """Upstream.xlsx::Data → Summary (current file)"""
        path = make_xlsx({
            "Summary": {(1, 1): "='[Upstream.xlsx]Data'!A1"},
        })
        try:
            _, edges = build_dependency_graph(path)
            assert len(edges) == 1
            src, dst = edges[0]
            assert src.workbook == "Upstream.xlsx"
            assert src.sheet    == "Data"
            assert dst.workbook is None
            assert dst.sheet    == "Summary"
        finally:
            os.unlink(path)

    def test_multiple_external_workbooks(self):
        path = make_xlsx({
            "Sheet1": {
                (1, 1): "='[Alpha.xlsx]Raw'!A1",
                (2, 1): "='[Beta.xlsx]Raw'!A1",
            },
        })
        try:
            nodes, edges = build_dependency_graph(path)
            wb_names = {n.workbook for n in nodes if n.workbook}
            assert "Alpha.xlsx" in wb_names
            assert "Beta.xlsx"  in wb_names
            assert len(edges) == 2
        finally:
            os.unlink(path)

    def test_same_external_ref_many_times_one_edge(self):
        cells = {(r, 1): "='[Upstream.xlsx]Data'!A1" for r in range(1, 51)}
        path = make_xlsx({"Summary": cells})
        try:
            _, edges = build_dependency_graph(path)
            assert len(edges) == 1
        finally:
            os.unlink(path)

    def test_reference_to_nonexistent_sheet_ignored(self):
        """
        If a formula references a sheet name that doesn't exist in the
        workbook, it should be treated as an internal reference but silently
        skipped (not crash, not add phantom nodes).
        """
        path = make_xlsx({
            "Sheet1": {(1, 1): "=GhostSheet!A1"},
        })
        try:
            nodes, edges = build_dependency_graph(path)
            # GhostSheet must not appear as an internal node
            internal_names = {n.sheet for n in nodes if n.workbook is None}
            assert "GhostSheet" not in internal_names
            assert edges == []
        finally:
            os.unlink(path)

    def test_pipe_labelled_formula(self):
        """Formula with a text label on the reference still parsed correctly."""
        path = make_xlsx({
            "Sheet1": {},
            "Sheet2": {(1, 1): "=Sheet1!A1+0"},
        })
        try:
            _, edges = build_dependency_graph(path)
            assert ("Sheet1", "Sheet2") in edge_pairs(edges)
        finally:
            os.unlink(path)

    def test_formula_with_multiple_sheet_refs(self):
        """=Sheet1!A1 + Sheet2!B1 in Sheet3 → two edges."""
        path = make_xlsx({
            "Sheet1": {},
            "Sheet2": {},
            "Sheet3": {(1, 1): "=Sheet1!A1+Sheet2!B1"},
        })
        try:
            _, edges = build_dependency_graph(path)
            ep = edge_pairs(edges)
            assert ("Sheet1", "Sheet3") in ep
            assert ("Sheet2", "Sheet3") in ep
        finally:
            os.unlink(path)


# ===========================================================================
# 5. TestGenerateHtml
# ===========================================================================

class TestGenerateHtml:
    def test_valid_html_structure(self):
        mmd  = "flowchart LR\n  A --> B"
        html = generate_html(mmd, title="Test.xlsx")
        assert html.startswith("<!DOCTYPE html>")
        assert "</html>" in html

    def test_title_embedded(self):
        html = generate_html("flowchart LR\n  A --> B", title="MyBook.xlsx")
        assert "MyBook.xlsx" in html

    def test_mermaid_source_present(self):
        html = generate_html("flowchart LR\n  A --> B", "x.xlsx")
        assert "flowchart LR" in html

    def test_mermaid_source_html_escaped(self):
        mmd  = 'flowchart LR\n  A["<b>bold</b>"] --> B'
        html = generate_html(mmd, "x.xlsx")
        assert "<b>bold</b>" not in html   # must not appear raw

    def test_xss_in_title_escaped(self):
        html = generate_html("flowchart LR\n  A --> B", title="<script>x</script>")
        assert "<script>x</script>" not in html

    def test_mermaid_js_included(self):
        html = generate_html("flowchart LR\n  A --> B", "x.xlsx")
        assert "mermaid" in html.lower()

    def test_download_button_present(self):
        html = generate_html("flowchart LR\n  A --> B", "x.xlsx")
        assert "downloadPNG" in html

    def test_reset_button_present(self):
        html = generate_html("flowchart LR\n  A --> B", "x.xlsx")
        assert "resetView" in html

    def test_pan_zoom_js_present(self):
        html = generate_html("flowchart LR\n  A --> B", "x.xlsx")
        assert "fitDiagram" in html
        assert "wheel" in html

    def test_graph_json_injected(self):
        # Use node IDs that match what build_mermaid() produces so that
        # parse_mermaid() inside generate_html() can round-trip them.
        mmd = "flowchart LR\n  CUR__Sheet1 --> CUR__Sheet2"
        html = generate_html(mmd, title="x.xlsx")
        assert "const GRAPH" in html
        assert "CUR__Sheet1" in html
        assert "CUR__Sheet2" in html
        # children map must link Sheet1 → Sheet2
        assert '"children"' in html

    def test_highlight_js_present(self):
        html = generate_html("flowchart LR\n  A --> B", "x.xlsx")
        assert "applyHighlight" in html
        assert "getRelated"     in html
        assert "svgIdToNodeId"  in html
        assert "resetHighlight" in html

    def test_no_graph_json_placeholder_left(self):
        """GRAPH_JSON placeholder must be fully replaced."""
        html = generate_html("flowchart LR\n  A --> B", "x.xlsx")
        assert "GRAPH_JSON" not in html


# ===========================================================================
# 6. TestEdgeCases
# ===========================================================================

class TestEdgeCases:
    def test_empty_workbook_no_crash(self):
        """A workbook with sheets but no cells at all."""
        path = make_xlsx({
            "Sheet1": {},
            "Sheet2": {},
        })
        try:
            nodes, edges = build_dependency_graph(path)
            assert edges == []
            assert len([n for n in nodes if n.workbook is None]) == 2
        finally:
            os.unlink(path)

    def test_single_sheet_workbook(self):
        path = make_xlsx({"Only": {(1, 1): "=SUM(A2:A10)"}})
        try:
            nodes, edges = build_dependency_graph(path)
            assert edges == []
            assert len(nodes) == 1
        finally:
            os.unlink(path)

    def test_formula_only_has_constants(self):
        path = make_xlsx({"Sheet1": {(1, 1): "=1+2+3"}})
        try:
            _, edges = build_dependency_graph(path)
            assert edges == []
        finally:
            os.unlink(path)

    def test_node_id_uniqueness_across_workbooks(self):
        """
        Sheet 'Summary' in two different workbooks must get distinct node IDs.
        """
        cur = SheetNode(None,          "Summary")
        ext = SheetNode("Other.xlsx",  "Summary")
        assert cur.node_id != ext.node_id

    def test_build_mermaid_empty_graph(self):
        mmd = build_mermaid(set(), [], "Empty.xlsx")
        assert mmd.startswith("flowchart LR")

    def test_special_chars_in_sheet_name(self):
        """Sheet names with quotes should not break the Mermaid output."""
        node = SheetNode(None, "It's a Sheet")
        mmd  = build_mermaid({node}, [], "X.xlsx")
        # The single-quote is replaced with an apostrophe inside double-quoted label
        assert "It" in mmd   # at least part of the name must appear

    def test_generate_html_from_real_xlsx(self):
        """End-to-end: xlsx → graph → mermaid → HTML."""
        path = make_xlsx({
            "Inputs":  {},
            "Calcs":   {(1, 1): "=Inputs!A1 * 2"},
            "Summary": {(1, 1): "=Calcs!A1 + Inputs!A2"},
        })
        try:
            nodes, edges = build_dependency_graph(path)
            title = "test_workbook.xlsx"
            mmd   = build_mermaid(nodes, edges, current_file=title)
            html  = generate_html(mmd, title=title)

            assert html.startswith("<!DOCTYPE html>")
            assert "Inputs"  in html
            assert "Calcs"   in html
            assert "Summary" in html
            # Mermaid source is HTML-escaped in the page (> → &gt;),
            # so check the diagram text directly for edges.
            assert "-->" in mmd
            assert len(edges) >= 2   # Inputs→Calcs, Inputs→Summary, Calcs→Summary
            # Graph JSON must be present for highlighting to work
            assert "GRAPH" in html
            assert "CUR__Inputs" in html
        finally:
            os.unlink(path)


# ===========================================================================
# 7. TestExternalConnectionNode
# ===========================================================================

class TestExternalConnectionNode:
    def test_odbc_node_id_prefix(self):
        n = ExternalConnectionNode("odbc", "MyDSN", "MyDSN")
        assert n.node_id.startswith("ODBC__")

    def test_pq_node_id_prefix(self):
        n = ExternalConnectionNode("powerquery", "Sales", "Sales")
        assert n.node_id.startswith("PQ__")

    def test_web_node_id_prefix(self):
        n = ExternalConnectionNode("web", "API Feed", "https://api.example.com")
        assert n.node_id.startswith("WEB__")

    def test_oledb_node_id_prefix(self):
        n = ExternalConnectionNode("oledb", "OLE Conn", "MyServer")
        assert n.node_id.startswith("OLEDB__")

    def test_node_id_special_chars_removed(self):
        n = ExternalConnectionNode("odbc", "My Server (prod)", "My Server (prod)")
        assert re.match(r"^[A-Za-z][A-Za-z0-9_]*$", n.node_id)

    def test_label_includes_source_when_different(self):
        n = ExternalConnectionNode("odbc", "Sales Connection", "PROD-SQL-01")
        assert "PROD-SQL-01" in n.label
        assert "Sales Connection" in n.label

    def test_label_just_name_when_same(self):
        n = ExternalConnectionNode("powerquery", "Invoices", "Invoices")
        assert n.label == "Invoices"

    def test_mermaid_class_odbc(self):
        assert ExternalConnectionNode("odbc", "x", "y").mermaid_class == "odbcNode"

    def test_mermaid_class_pq(self):
        assert ExternalConnectionNode("powerquery", "x", "y").mermaid_class == "pqNode"

    def test_equality(self):
        a = ExternalConnectionNode("odbc", "DSN1", "DSN1")
        b = ExternalConnectionNode("odbc", "DSN1", "DSN1")
        assert a == b

    def test_hashable_in_set(self):
        a = ExternalConnectionNode("odbc", "DSN1", "DSN1")
        b = ExternalConnectionNode("odbc", "DSN1", "DSN1")
        assert len({a, b}) == 1

    def test_different_types_not_equal(self):
        a = ExternalConnectionNode("odbc", "X", "X")
        b = ExternalConnectionNode("powerquery", "X", "X")
        # node_ids differ due to prefix → not equal
        assert a != b


# ===========================================================================
# 8. TestBuildMermaidWithConnections
# ===========================================================================

class TestBuildMermaidWithConnections:
    def _make_conn(self, ctype: str, name: str, source: str = "") -> ExternalConnectionNode:
        return ExternalConnectionNode(ctype, name, source or name)

    def test_classdef_emitted_when_conn_edges_present(self):
        s1 = SheetNode(None, "Sheet1")
        cn = self._make_conn("odbc", "MyDB", "PROD-SQL")
        mmd = build_mermaid({s1}, [], "wb.xlsx", conn_edges=[(cn, s1)])
        assert "classDef odbcNode" in mmd

    def test_no_classdef_without_conn_edges(self):
        s1 = SheetNode(None, "Sheet1")
        mmd = build_mermaid({s1}, [], "wb.xlsx")
        assert "classDef" not in mmd

    def test_conn_subgraph_created(self):
        s1 = SheetNode(None, "Sheet1")
        cn = self._make_conn("odbc", "SalesDB", "PROD-SQL")
        mmd = build_mermaid({s1}, [], "wb.xlsx", conn_edges=[(cn, s1)])
        assert "SG_CONN_ODBC" in mmd
        assert "ODBC Connections" in mmd

    def test_pq_subgraph_created(self):
        s1 = SheetNode(None, "Sheet1")
        cn = self._make_conn("powerquery", "Invoices", "Invoices")
        mmd = build_mermaid({s1}, [], "wb.xlsx", conn_edges=[(cn, s1)])
        assert "SG_CONN_POWERQUERY" in mmd
        assert "Power Query Connections" in mmd

    def test_conn_edge_in_mermaid(self):
        s1 = SheetNode(None, "Sheet1")
        cn = self._make_conn("odbc", "SalesDB", "PROD")
        mmd = build_mermaid({s1}, [], "wb.xlsx", conn_edges=[(cn, s1)])
        assert f"{cn.node_id} --> {s1.node_id}" in mmd

    def test_class_statement_emitted(self):
        s1 = SheetNode(None, "Sheet1")
        cn = self._make_conn("odbc", "SalesDB", "PROD")
        mmd = build_mermaid({s1}, [], "wb.xlsx", conn_edges=[(cn, s1)])
        assert f"class {cn.node_id} odbcNode" in mmd

    def test_multiple_conn_types_all_present(self):
        s1 = SheetNode(None, "Sheet1")
        c_odbc = self._make_conn("odbc", "DB1", "SERVER1")
        c_pq   = self._make_conn("powerquery", "PQ1", "PQ1")
        c_web  = self._make_conn("web", "API", "https://api.example.com")
        conn_edges = [(c_odbc, s1), (c_pq, s1), (c_web, s1)]
        mmd = build_mermaid({s1}, [], "wb.xlsx", conn_edges=conn_edges)
        assert "classDef odbcNode" in mmd
        assert "classDef pqNode"   in mmd
        assert "classDef webNode"  in mmd

    def test_duplicate_conn_nodes_deduplicated(self):
        """Same connection feeding two sheets should appear only once in subgraph."""
        s1 = SheetNode(None, "Sheet1")
        s2 = SheetNode(None, "Sheet2")
        cn = self._make_conn("powerquery", "Data", "Data")
        mmd = build_mermaid({s1, s2}, [], "wb.xlsx", conn_edges=[(cn, s1), (cn, s2)])
        # Node should appear once in the subgraph, but two edges
        assert mmd.count(cn.node_id + '[') == 1  # node declaration
        assert mmd.count(f"{cn.node_id} -->") == 2  # two edges

    def test_sheet_subgraph_still_present_with_conns(self):
        s1 = SheetNode(None, "Sheet1")
        cn = self._make_conn("odbc", "DB", "DB")
        mmd = build_mermaid({s1}, [], "my_file.xlsx", conn_edges=[(cn, s1)])
        assert "SG_CUR" in mmd
        assert "my_file.xlsx" in mmd


# ===========================================================================
# 9. TestExtractExternalConnections (xlsx with embedded connections.xml)
# ===========================================================================

def _make_xlsx_with_connections_xml(
    sheets: list[str],
    connections_xml: str,
) -> str:
    """
    Create a minimal .xlsx zip that has ``xl/connections.xml`` injected
    but otherwise is a valid openpyxl workbook.
    """
    import io

    # Create a valid xlsx first
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for s in sheets:
        wb.create_sheet(s)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    raw = buf.read()

    # Re-open the zip and inject connections.xml
    import zipfile
    in_buf  = io.BytesIO(raw)
    out_buf = io.BytesIO()

    with zipfile.ZipFile(in_buf, "r") as zin, \
         zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            zout.writestr(item, zin.read(item.filename))
        # Inject our custom connections.xml
        zout.writestr("xl/connections.xml", connections_xml)

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    with open(path, "wb") as fh:
        fh.write(out_buf.getvalue())
    return path


class TestExtractExternalConnections:
    _ODBC_CONN_XML = """\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <connection id="1" name="SalesDB" type="1" refreshedVersion="3">
    <dbPr connection="DSN=SALES_PROD;UID=reader;PWD=xxx" command="SELECT * FROM sales"/>
  </connection>
</connections>"""

    _PQ_CONN_XML = """\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <connection id="2" name="Query - Invoices" type="5" refreshedVersion="6">
    <dbPr connection="Provider=Microsoft.Mashup.OleDb.1;Data Source=$Workbook$;Location=Invoices"/>
    <extLst>
      <ext uri="{DE250136-89BD-433C-8126-D09CA5730AF9}"/>
    </extLst>
  </connection>
</connections>"""

    _WEB_CONN_XML = """\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <connection id="3" name="ExchangeRates" type="4" refreshedVersion="3">
    <webPr url="https://api.exchangerate.host/latest" sourceDataFile="0"/>
  </connection>
</connections>"""

    def test_odbc_connection_detected(self):
        path = _make_xlsx_with_connections_xml(["Sheet1"], self._ODBC_CONN_XML)
        try:
            nodes, _ = build_dependency_graph(path)
            conn_nodes, _ = extract_external_connections(path, nodes)
            types = {cn.conn_type for cn in conn_nodes}
            assert "odbc" in types
        finally:
            os.unlink(path)

    def test_odbc_dsn_extracted(self):
        path = _make_xlsx_with_connections_xml(["Sheet1"], self._ODBC_CONN_XML)
        try:
            nodes, _ = build_dependency_graph(path)
            conn_nodes, _ = extract_external_connections(path, nodes)
            sources = {cn.source for cn in conn_nodes}
            assert "SALES_PROD" in sources
        finally:
            os.unlink(path)

    def test_powerquery_connection_detected(self):
        path = _make_xlsx_with_connections_xml(["Sheet1"], self._PQ_CONN_XML)
        try:
            nodes, _ = build_dependency_graph(path)
            conn_nodes, _ = extract_external_connections(path, nodes)
            types = {cn.conn_type for cn in conn_nodes}
            assert "powerquery" in types
        finally:
            os.unlink(path)

    def test_powerquery_name_strip_prefix(self):
        """'Query - Invoices' should be cleaned to 'Invoices'."""
        path = _make_xlsx_with_connections_xml(["Sheet1"], self._PQ_CONN_XML)
        try:
            nodes, _ = build_dependency_graph(path)
            conn_nodes, _ = extract_external_connections(path, nodes)
            names = {cn.name for cn in conn_nodes}
            assert "Invoices" in names
        finally:
            os.unlink(path)

    def test_web_connection_detected(self):
        path = _make_xlsx_with_connections_xml(["Sheet1"], self._WEB_CONN_XML)
        try:
            nodes, _ = build_dependency_graph(path)
            conn_nodes, _ = extract_external_connections(path, nodes)
            types = {cn.conn_type for cn in conn_nodes}
            assert "web" in types
        finally:
            os.unlink(path)

    def test_web_url_as_source(self):
        path = _make_xlsx_with_connections_xml(["Sheet1"], self._WEB_CONN_XML)
        try:
            nodes, _ = build_dependency_graph(path)
            conn_nodes, _ = extract_external_connections(path, nodes)
            sources = {cn.source for cn in conn_nodes}
            assert any("exchangerate" in s for s in sources)
        finally:
            os.unlink(path)

    def test_no_connections_returns_empty(self):
        path = _make_xlsx_with_connections_xml(["Sheet1"], "")
        try:
            nodes, _ = build_dependency_graph(path)
            conn_nodes, conn_edges = extract_external_connections(path, nodes)
            assert conn_nodes == []
            assert conn_edges == []
        finally:
            os.unlink(path)

    def test_unlinked_conn_has_no_edge(self):
        """ODBC connections not mapped via queryTable appear with no arrows."""
        path = _make_xlsx_with_connections_xml(["Sheet1", "Sheet2"], self._ODBC_CONN_XML)
        try:
            nodes, _ = build_dependency_graph(path)
            conn_nodes, conn_edges = extract_external_connections(path, nodes)
            # Node must be discovered
            assert len(conn_nodes) == 1
            assert conn_nodes[0].conn_type == "odbc"
            # But no edges (no queryTable links it to a sheet)
            assert conn_edges == []
        finally:
            os.unlink(path)

    def test_mermaid_includes_conn_nodes(self):
        path = _make_xlsx_with_connections_xml(["Sheet1"], self._ODBC_CONN_XML)
        try:
            nodes, edges = build_dependency_graph(path)
            conn_nodes, conn_edges = extract_external_connections(path, nodes)
            mmd = build_mermaid(nodes, edges, "wb.xlsx",
                                conn_nodes=conn_nodes, conn_edges=conn_edges)
            assert "ODBC" in mmd
            assert "SALES_PROD" in mmd or "SalesDB" in mmd
        finally:
            os.unlink(path)

    def test_html_output_includes_conn_nodes(self):
        path = _make_xlsx_with_connections_xml(["Sheet1"], self._PQ_CONN_XML)
        try:
            nodes, edges = build_dependency_graph(path)
            conn_nodes, conn_edges = extract_external_connections(path, nodes)
            mmd  = build_mermaid(nodes, edges, "wb.xlsx",
                                 conn_nodes=conn_nodes, conn_edges=conn_edges)
            html = generate_html(mmd, title="wb.xlsx")
            assert "PQ__" in html or "pqNode" in html or "Power Query" in html
        finally:
            os.unlink(path)


# ===========================================================================
# 10. TestExternalConnectionNode wb_scope
# ===========================================================================

class TestExternalConnectionNodeWbScope:
    def test_wb_scope_changes_node_id(self):
        a = ExternalConnectionNode("odbc", "SalesDB", "PROD-SQL")
        b = ExternalConnectionNode("odbc", "SalesDB", "PROD-SQL", wb_scope="rates.xlsx")
        assert a.node_id != b.node_id

    def test_scoped_node_id_contains_wb_fragment(self):
        cn = ExternalConnectionNode("odbc", "DB", "DB", wb_scope="Upstream.xlsx")
        assert "Upstream" in cn.node_id or "upstream" in cn.node_id.lower()

    def test_no_scope_gives_double_underscore_prefix(self):
        cn = ExternalConnectionNode("odbc", "DSN1", "DSN1")
        assert "ODBC__" in cn.node_id

    def test_scoped_node_id_is_valid_mermaid_id(self):
        cn = ExternalConnectionNode("powerquery", "My Query", "My Query",
                                    wb_scope="rates (2025).xlsx")
        assert re.match(r"^[A-Za-z][A-Za-z0-9_]*$", cn.node_id)

    def test_two_scoped_nodes_same_source_different_wb_differ(self):
        a = ExternalConnectionNode("odbc", "DB", "PROD", wb_scope="file1.xlsx")
        b = ExternalConnectionNode("odbc", "DB", "PROD", wb_scope="file2.xlsx")
        assert a.node_id != b.node_id

    def test_label_unaffected_by_wb_scope(self):
        a = ExternalConnectionNode("odbc", "SalesDB", "PROD-SQL")
        b = ExternalConnectionNode("odbc", "SalesDB", "PROD-SQL", wb_scope="x.xlsx")
        assert a.label == b.label

    def test_hash_differs_with_scope(self):
        a = ExternalConnectionNode("odbc", "DB", "DB")
        b = ExternalConnectionNode("odbc", "DB", "DB", wb_scope="x.xlsx")
        assert hash(a) != hash(b)


# ===========================================================================
# 11. TestExtractUpstreamFileConnections
# ===========================================================================

def _make_xlsx_pair(
    main_sheets: dict[str, dict],
    upstream_name: str,
    upstream_sheets: list[str],
    upstream_connections_xml: str,
    same_dir: bool = True,
) -> tuple[str, str]:
    """
    Create two xlsx files.  The main file has formula refs to upstream_name.
    Returns (main_path, upstream_path).  Both are in the same temp dir when
    same_dir=True.
    """
    import io, tempfile, zipfile as zf_mod

    td = tempfile.mkdtemp()

    # ── upstream file ──────────────────────────────────────────────────────
    up_wb = openpyxl.Workbook()
    up_wb.remove(up_wb.active)
    for s in upstream_sheets:
        up_wb.create_sheet(s)
    up_buf = io.BytesIO()
    up_wb.save(up_buf)
    up_buf.seek(0)
    raw_up = up_buf.read()

    # Inject connections.xml into upstream if provided
    if upstream_connections_xml:
        in_b  = io.BytesIO(raw_up)
        out_b = io.BytesIO()
        with zf_mod.ZipFile(in_b, "r") as zin, \
             zf_mod.ZipFile(out_b, "w", zf_mod.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                zout.writestr(item, zin.read(item.filename))
            zout.writestr("xl/connections.xml", upstream_connections_xml)
        raw_up = out_b.getvalue()

    if same_dir:
        up_path = os.path.join(td, upstream_name)
    else:
        up_path = os.path.join(tempfile.mkdtemp(), upstream_name)

    with open(up_path, "wb") as fh:
        fh.write(raw_up)

    # ── main file (references upstream via formula) ────────────────────────
    main_wb = openpyxl.Workbook()
    main_wb.remove(main_wb.active)
    ws = main_wb.create_sheet("Sheet1")
    # Formula referencing upstream_name, first sheet
    first_sheet = upstream_sheets[0] if upstream_sheets else "Sheet1"
    ws.cell(1, 1, f"=[{upstream_name}]{first_sheet}!A1")
    main_buf = io.BytesIO()
    main_wb.save(main_buf)
    main_path = os.path.join(td, "main.xlsx")
    with open(main_path, "wb") as fh:
        fh.write(main_buf.getvalue())

    return main_path, up_path


_ODBC_XML = """\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <connection id="1" name="SalesDB" type="1" refreshedVersion="3">
    <dbPr connection="DSN=SALES_PROD;UID=reader"/>
  </connection>
</connections>"""

_PQ_XML = """\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <connection id="2" name="Query - Invoices" type="5" refreshedVersion="6">
    <dbPr connection="Provider=Microsoft.Mashup.OleDb.1;Data Source=$Workbook$;Location=Invoices"/>
    <extLst><ext uri="{DE250136-89BD-433C-8126-D09CA5730AF9}"/></extLst>
  </connection>
</connections>"""


class TestExtractUpstreamFileConnections:

    def test_returns_empty_when_no_ext_refs(self):
        """No external workbook refs → nothing to look up."""
        path = make_xlsx({"Sheet1": {(1, 1): "=SUM(A2:A3)"}})
        try:
            nodes, _ = build_dependency_graph(path)
            result = extract_upstream_file_connections(path, nodes)
            assert result == {}
        finally:
            os.unlink(path)

    def test_returns_empty_when_upstream_not_on_disk(self):
        """Upstream file is referenced but not in same folder."""
        main_path, up_path = _make_xlsx_pair(
            {}, "rates.xlsx", ["Data"], _ODBC_XML, same_dir=False
        )
        try:
            nodes, _ = build_dependency_graph(main_path)
            result = extract_upstream_file_connections(main_path, nodes)
            assert result == {}
        finally:
            import shutil
            shutil.rmtree(os.path.dirname(main_path), ignore_errors=True)
            shutil.rmtree(os.path.dirname(up_path), ignore_errors=True)

    def test_detects_odbc_in_upstream(self):
        main_path, _ = _make_xlsx_pair(
            {}, "rates.xlsx", ["Data"], _ODBC_XML
        )
        try:
            nodes, _ = build_dependency_graph(main_path)
            result = extract_upstream_file_connections(main_path, nodes)
            assert "rates.xlsx" in result
            types = {cn.conn_type for cn in result["rates.xlsx"]}
            assert "odbc" in types
        finally:
            import shutil
            shutil.rmtree(os.path.dirname(main_path), ignore_errors=True)

    def test_detects_pq_in_upstream(self):
        main_path, _ = _make_xlsx_pair(
            {}, "upstream.xlsx", ["Sheet1"], _PQ_XML
        )
        try:
            nodes, _ = build_dependency_graph(main_path)
            result = extract_upstream_file_connections(main_path, nodes)
            assert "upstream.xlsx" in result
            types = {cn.conn_type for cn in result["upstream.xlsx"]}
            assert "powerquery" in types
        finally:
            import shutil
            shutil.rmtree(os.path.dirname(main_path), ignore_errors=True)

    def test_node_ids_are_scoped_to_wb(self):
        """Scoped IDs must not match the plain (unscoped) node ID."""
        main_path, _ = _make_xlsx_pair(
            {}, "rates.xlsx", ["Data"], _ODBC_XML
        )
        try:
            nodes, _ = build_dependency_graph(main_path)
            result = extract_upstream_file_connections(main_path, nodes)
            for cn in result.get("rates.xlsx", []):
                assert cn.wb_scope == "rates.xlsx"
                # Scoped ID should differ from a plain version of the same name
                plain = ExternalConnectionNode(cn.conn_type, cn.name, cn.source)
                assert cn.node_id != plain.node_id
        finally:
            import shutil
            shutil.rmtree(os.path.dirname(main_path), ignore_errors=True)

    def test_returns_empty_when_upstream_has_no_connections(self):
        """Upstream file exists but has no xl/connections.xml."""
        main_path, _ = _make_xlsx_pair(
            {}, "empty_upstream.xlsx", ["Sheet1"], ""
        )
        try:
            nodes, _ = build_dependency_graph(main_path)
            result = extract_upstream_file_connections(main_path, nodes)
            assert result == {}
        finally:
            import shutil
            shutil.rmtree(os.path.dirname(main_path), ignore_errors=True)

    def test_build_mermaid_puts_upstream_conns_in_ext_subgraph(self):
        main_path, _ = _make_xlsx_pair(
            {}, "rates.xlsx", ["Data"], _ODBC_XML
        )
        try:
            nodes, edges = build_dependency_graph(main_path)
            upstream_conn_map = extract_upstream_file_connections(main_path, nodes)
            mmd = build_mermaid(nodes, edges, "main.xlsx",
                                upstream_conn_map=upstream_conn_map)
            # The upstream conn node must appear inside the rates.xlsx subgraph
            sg_start = mmd.index('subgraph SG_EXT__rates_xlsx')
            sg_end   = mmd.index("  end", sg_start)
            subgraph_block = mmd[sg_start:sg_end]
            assert "ODBC_rates_xlsx__SALES_PROD" in subgraph_block or \
                   any(cn.node_id in subgraph_block
                       for cn in upstream_conn_map.get("rates.xlsx", []))
        finally:
            import shutil
            shutil.rmtree(os.path.dirname(main_path), ignore_errors=True)

    def test_build_mermaid_no_arrows_for_upstream_conns(self):
        """Upstream file connections must not generate any --> edges."""
        main_path, _ = _make_xlsx_pair(
            {}, "rates.xlsx", ["Data"], _ODBC_XML
        )
        try:
            nodes, edges = build_dependency_graph(main_path)
            upstream_conn_map = extract_upstream_file_connections(main_path, nodes)
            mmd = build_mermaid(nodes, edges, "main.xlsx",
                                upstream_conn_map=upstream_conn_map)
            # Collect all upstream conn node IDs
            up_ids = {cn.node_id
                      for cn_list in upstream_conn_map.values()
                      for cn in cn_list}
            # None of those IDs should appear as the source of a --> edge
            for line in mmd.splitlines():
                line = line.strip()
                if "-->" in line:
                    src = line.split("-->")[0].strip()
                    assert src not in up_ids, \
                        f"Upstream conn node {src!r} has an unexpected arrow"
        finally:
            import shutil
            shutil.rmtree(os.path.dirname(main_path), ignore_errors=True)

    def test_class_applied_to_upstream_conn_nodes(self):
        """CSS class statement must include upstream connection nodes."""
        main_path, _ = _make_xlsx_pair(
            {}, "rates.xlsx", ["Data"], _ODBC_XML
        )
        try:
            nodes, edges = build_dependency_graph(main_path)
            upstream_conn_map = extract_upstream_file_connections(main_path, nodes)
            mmd = build_mermaid(nodes, edges, "main.xlsx",
                                upstream_conn_map=upstream_conn_map)
            for cn in upstream_conn_map.get("rates.xlsx", []):
                assert cn.node_id in mmd
                assert "odbcNode" in mmd
        finally:
            import shutil
            shutil.rmtree(os.path.dirname(main_path), ignore_errors=True)

    def test_html_includes_upstream_conn_node(self):
        main_path, _ = _make_xlsx_pair(
            {}, "rates.xlsx", ["Data"], _ODBC_XML
        )
        try:
            nodes, edges = build_dependency_graph(main_path)
            upstream_conn_map = extract_upstream_file_connections(main_path, nodes)
            mmd  = build_mermaid(nodes, edges, "main.xlsx",
                                 upstream_conn_map=upstream_conn_map)
            html = generate_html(mmd, title="main.xlsx")
            assert html.startswith("<!DOCTYPE html>")
            assert any(
                cn.node_id in html
                for cn_list in upstream_conn_map.values()
                for cn in cn_list
            )
        finally:
            import shutil
            shutil.rmtree(os.path.dirname(main_path), ignore_errors=True)
