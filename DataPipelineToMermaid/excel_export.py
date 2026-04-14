"""Convert a PipelineLineage JSON into a multi-sheet Excel workbook.

Sheets
------
1. Overview       — pipeline metadata & counts
2. Source Tables   — schema, table, column, type, description
3. Target Tables   — same layout as Source Tables
4. Column Lineage  — target ← source(s), transformation, type, steps
5. Components      — CTEs, procs, transforms
6. Data Flow       — edge list (from → to, columns)

artifact_mode (from config.yaml)
---------------------------------
regular : one row per target column (default)
long    : one row per file in the lineage chain; key columns repeat on every
          row; rows are ordered upstream-first (execution order).
wide    : one row per target column; files expand rightward as hop clusters
          prefixed hop_1_, hop_2_, …; each cluster color-coded.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import NamedTuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import ColumnLineage, PipelineLineage

# ── Style constants ─────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
_HEADER_FILL = PatternFill(
    start_color="2F5496", end_color="2F5496", fill_type="solid"
)
_HEADER_ALIGN = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)
_ALT_ROW_FILL = PatternFill(
    start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_WRAP = Alignment(vertical="top", wrap_text=True)

# Cycling color palette for wide-mode hop clusters (header fill / data tint)
_HOP_HEADER_COLORS = [
    "2F5496",  # blue
    "375623",  # green
    "843C0C",  # burnt orange
    "4B0082",  # indigo
    "005F60",  # teal
    "7B0000",  # dark red
    "665202",  # dark gold
    "1C3A5E",  # navy
]
_HOP_DATA_COLORS = [
    "D6E4F0",  # light blue
    "E2EFDA",  # light green
    "FCE4D6",  # light orange
    "E8D5F5",  # light purple
    "D5F0EE",  # light teal
    "FCDBD9",  # light red
    "FFF2CC",  # light gold
    "D9E1F2",  # light navy
]
_COLS_PER_HOP = 6  # file, raw_source_tables, raw_source_columns, transformation_step, type, notes


# ── Long-mode helpers ───────────────────────────────────────────────

_STEP_RE = re.compile(r"→\s+\[(?P<file>[^\]]+)\]\s+(?P<expr>.+)")


class _FileStep(NamedTuple):
    file: str
    expr: str  # one expression from the transformation chain


def _parse_chain(transformation: str) -> list[_FileStep]:
    """Parse a transformation chain into ordered (file, expr) pairs.

    Input format::

        SOURCES: raw.t.col1, raw.t.col2
        → [01_raw_orders.sql] t.gross_value AS raw_amount
        → [04_enrich_orders.mp] raw_amount * ... AS discounted_amount

    Returns a list of ``_FileStep`` in chain order (upstream → downstream).
    Lines that don't match the ``→ [file] expr`` pattern are skipped.
    """
    steps: list[_FileStep] = []
    for line in transformation.splitlines():
        m = _STEP_RE.match(line.strip())
        if m:
            steps.append(_FileStep(file=m.group("file"), expr=m.group("expr")))
    return steps


def _group_chain_by_file(steps: list[_FileStep]) -> list[tuple[str, str]]:
    """Group chain steps by file, preserving first-appearance order.

    When a file contributes multiple expressions, they are joined with
    a newline so the row's transformation cell is self-contained.

    Returns ``[(filename, combined_exprs), ...]`` ordered upstream-first.
    """
    seen: dict[str, list[str]] = {}   # preserves insertion order (Python 3.7+)
    for step in steps:
        seen.setdefault(step.file, []).append(step.expr)
    return [(fname, "\n".join(exprs)) for fname, exprs in seen.items()]


# ── Helpers ─────────────────────────────────────────────────────────


def _style_header(ws: Worksheet, ncols: int) -> None:
    """Apply header styling to row 1."""
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _auto_width(ws: Worksheet, *, min_w: int = 14, max_w: int = 55) -> None:
    """Set column widths based on content length."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = max(
            (len(str(c.value)) if c.value else 0 for c in col_cells),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(
            max(length + 3, min_w), max_w
        )


def _stripe_rows(ws: Worksheet, ncols: int) -> None:
    """Apply alternating row colours and borders from row 2 onward."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=ncols), 2):
        for cell in row:
            cell.border = _THIN_BORDER
            cell.alignment = _WRAP
            if row_idx % 2 == 0:
                cell.fill = _ALT_ROW_FILL


# ── Column Lineage sheet writers ────────────────────────────────────


def _write_column_lineage_regular(
    ws: Worksheet,
    column_lineage: list[ColumnLineage],
) -> None:
    """Write one row per target column (regular mode)."""
    has_steps = any(cl.intermediate_steps for cl in column_lineage)

    hdr = [
        "Target Table",
        "Target Column",
        "Source Table(s)",
        "Source Column(s)",
        "Transformation (SQL)",
        "Type",
        "Source File(s)",
        "Target File",
        "Notes",
    ]
    if has_steps:
        hdr.insert(-1, "Intermediate Steps")
    ws.append(hdr)
    _style_header(ws, len(hdr))

    for cl in column_lineage:
        src_tables = "\n".join(ref.source_table for ref in cl.source_refs)
        src_cols   = "\n".join(ref.source_column for ref in cl.source_refs)
        src_files  = ", ".join(cl.source_filenames) if cl.source_filenames else cl.filename
        row = [
            cl.target_table,
            cl.target_column,
            src_tables,
            src_cols,
            cl.transformation,
            cl.transformation_type,
            src_files,
            cl.filename,
            cl.notes,
        ]
        if has_steps:
            steps_str = (
                " → ".join(
                    f"[{s.component_name}] {s.expression} → {s.output_column}"
                    for s in cl.intermediate_steps
                )
                if cl.intermediate_steps else ""
            )
            row.insert(-1, steps_str)
        ws.append(row)


def _write_column_lineage_long(
    ws: Worksheet,
    column_lineage: list[ColumnLineage],
) -> None:
    """Write one row per file per target column (long mode).

    Each target column expands to N rows — one per file in its lineage chain,
    ordered upstream-first.  All key columns repeat on every row so that a
    downstream LLM can interpret each row without resolving blank cells.

    Columns
    -------
    Target Table | Target Column | Step # | File | Transformation Step |
    Type | Raw Source Tables | Raw Source Columns | Notes
    """
    hdr = [
        "Target Table",
        "Target Column",
        "Step #",
        "File",
        "Transformation Step",
        "Type",
        "Raw Source Tables",
        "Raw Source Columns",
        "Notes",
    ]
    ws.append(hdr)
    _style_header(ws, len(hdr))

    for cl in column_lineage:
        raw_tables = "\n".join(ref.source_table for ref in cl.source_refs)
        raw_cols   = "\n".join(ref.source_column for ref in cl.source_refs)

        # Parse the transformation chain into per-file groups
        chain_steps = _parse_chain(cl.transformation)
        file_groups = _group_chain_by_file(chain_steps)

        if not file_groups:
            # No chain found (e.g. not-found column) — emit one placeholder row
            ws.append([
                cl.target_table,
                cl.target_column,
                1,
                cl.filename or "",
                cl.transformation,
                cl.transformation_type,
                raw_tables,
                raw_cols,
                cl.notes,
            ])
            continue

        for step_num, (fname, expr) in enumerate(file_groups, start=1):
            ws.append([
                cl.target_table,
                cl.target_column,
                step_num,
                fname,
                expr,
                cl.transformation_type,   # repeats; applies to final output
                raw_tables,               # repeats for LLM readability
                raw_cols,                 # repeats for LLM readability
                cl.notes,
            ])


def _write_column_lineage_wide(
    ws: Worksheet,
    column_lineage: list[ColumnLineage],
) -> None:
    """Write one row per target column, files expanding rightward as hop clusters (wide mode).

    Each hop cluster has exactly ``_COLS_PER_HOP`` columns:
      hop_n_file | hop_n_raw_source_tables | hop_n_raw_source_columns |
      hop_n_transformation_step | hop_n_type | hop_n_notes

    Raw sources populate only hop_1 (blank for later hops).
    Type repeats every hop for LLM readability.
    Notes appear only on the final hop.
    Shorter columns leave trailing hop cluster cells blank.
    Header cells are color-coded per hop cluster (cycling palette).
    Data cells receive a light tint of the same palette color.
    """
    # ── Parse chains and determine max hops ────────────────────────
    all_groups: list[list[tuple[str, str]]] = []
    for cl in column_lineage:
        chain = _parse_chain(cl.transformation)
        all_groups.append(_group_chain_by_file(chain))

    max_hops = max((len(g) for g in all_groups), default=1)

    # ── Build and write header row ──────────────────────────────────
    base_hdr = ["Target Table", "Target Column"]
    hop_hdr_cols: list[list[str]] = [
        [
            f"hop_{n}_file",
            f"hop_{n}_raw_source_tables",
            f"hop_{n}_raw_source_columns",
            f"hop_{n}_transformation_step",
            f"hop_{n}_type",
            f"hop_{n}_notes",
        ]
        for n in range(1, max_hops + 1)
    ]
    hdr = base_hdr + [col for hop in hop_hdr_cols for col in hop]
    ws.append(hdr)

    # Style base header columns
    for c in range(1, len(base_hdr) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # Style hop cluster header columns (cycling colors)
    for hop_idx in range(max_hops):
        hdr_color = _HOP_HEADER_COLORS[hop_idx % len(_HOP_HEADER_COLORS)]
        fill = PatternFill(start_color=hdr_color, end_color=hdr_color, fill_type="solid")
        start_col = len(base_hdr) + 1 + hop_idx * _COLS_PER_HOP
        for c_offset in range(_COLS_PER_HOP):
            cell = ws.cell(row=1, column=start_col + c_offset)
            cell.font = _HEADER_FONT
            cell.fill = fill
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER

    # ── Write data rows ─────────────────────────────────────────────
    for row_idx, (cl, groups) in enumerate(zip(column_lineage, all_groups), start=2):
        raw_tables = "\n".join(ref.source_table for ref in cl.source_refs)
        raw_cols   = "\n".join(ref.source_column for ref in cl.source_refs)

        row: list = [cl.target_table, cl.target_column]
        for hop_idx in range(max_hops):
            if hop_idx < len(groups):
                fname, expr = groups[hop_idx]
                is_last_hop = hop_idx == len(groups) - 1
                row.extend([
                    fname,
                    raw_tables if hop_idx == 0 else "",   # raw sources only in hop_1
                    raw_cols   if hop_idx == 0 else "",
                    expr,
                    cl.transformation_type,               # repeat every hop for LLM
                    cl.notes if is_last_hop else "",      # notes on final hop only
                ])
            else:
                row.extend([""] * _COLS_PER_HOP)         # blank trailing hops

        ws.append(row)

        # Style base columns
        for c in range(1, len(base_hdr) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.alignment = _WRAP
            cell.border = _THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = _ALT_ROW_FILL

        # Style hop cluster columns with light tint
        for hop_idx in range(max_hops):
            data_color = _HOP_DATA_COLORS[hop_idx % len(_HOP_DATA_COLORS)]
            fill = PatternFill(start_color=data_color, end_color=data_color, fill_type="solid")
            start_col = len(base_hdr) + 1 + hop_idx * _COLS_PER_HOP
            for c_offset in range(_COLS_PER_HOP):
                cell = ws.cell(row=row_idx, column=start_col + c_offset)
                cell.fill = fill
                cell.alignment = _WRAP
                cell.border = _THIN_BORDER


# ── Main export function ───────────────────────────────────────────


def lineage_to_excel(
    lineage: PipelineLineage,
    output_path: str | Path,
    *,
    mode: str | None = None,
) -> Path:
    """Write *lineage* to a ``.xlsx`` workbook at *output_path*.

    Parameters
    ----------
    lineage:
        The pipeline lineage to export.
    output_path:
        Destination ``.xlsx`` path.
    mode:
        ``'regular'`` (default) or ``'long'``.  If *None*, the value is
        read from ``config.yaml`` via :func:`~.config.get_artifact_mode`.

    Returns the resolved output path.
    """
    if mode is None:
        try:
            from .config import get_artifact_mode
            mode = get_artifact_mode()
        except Exception:
            mode = "regular"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Overview ───────────────────────────────────────────

    ws = wb.active
    ws.title = "Overview"
    headers = ["Property", "Value"]
    ws.append(headers)
    _style_header(ws, len(headers))

    ws.append(["Pipeline Name", lineage.pipeline_name])
    ws.append(["Pipeline Type", lineage.pipeline_type])
    ws.append(["Source File", lineage.source_file])
    ws.append(["Description", lineage.description])
    ws.append([""])
    ws.append(["Source Tables", len(lineage.sources)])
    ws.append(["Target Tables", len(lineage.targets)])
    ws.append(["Components (CTEs / Procs / …)", len(lineage.components)])
    ws.append(["Column Lineage Mappings", len(lineage.column_lineage)])
    ws.append(["Data Flow Edges", len(lineage.data_flow_edges)])
    _stripe_rows(ws, len(headers))
    _auto_width(ws)

    # ── Sheet 2: Source Tables ──────────────────────────────────────

    ws_src = wb.create_sheet("Source Tables")
    src_hdr = ["Schema", "Table", "Column", "Data Type", "Description"]
    ws_src.append(src_hdr)
    _style_header(ws_src, len(src_hdr))

    for tbl in lineage.sources:
        if tbl.columns:
            for col in tbl.columns:
                ws_src.append([
                    tbl.schema_name,
                    tbl.table_name,
                    col.name,
                    col.data_type,
                    col.description,
                ])
        else:
            ws_src.append([tbl.schema_name, tbl.table_name, "", "", ""])
    _stripe_rows(ws_src, len(src_hdr))
    _auto_width(ws_src)

    # ── Sheet 3: Target Tables ──────────────────────────────────────

    ws_tgt = wb.create_sheet("Target Tables")
    tgt_hdr = ["Schema", "Table", "Column", "Data Type", "Description"]
    ws_tgt.append(tgt_hdr)
    _style_header(ws_tgt, len(tgt_hdr))

    for tbl in lineage.targets:
        if tbl.columns:
            for col in tbl.columns:
                ws_tgt.append([
                    tbl.schema_name,
                    tbl.table_name,
                    col.name,
                    col.data_type,
                    col.description,
                ])
        else:
            ws_tgt.append([tbl.schema_name, tbl.table_name, "", "", ""])
    _stripe_rows(ws_tgt, len(tgt_hdr))
    _auto_width(ws_tgt)

    # ── Sheet 4: Column Lineage ─────────────────────────────────────

    ws_lin = wb.create_sheet("Column Lineage")

    if mode == "long":
        _write_column_lineage_long(ws_lin, lineage.column_lineage)
        _stripe_rows(ws_lin, ws_lin.max_column)
    elif mode == "wide":
        _write_column_lineage_wide(ws_lin, lineage.column_lineage)
        # wide mode applies its own per-hop coloring; skip generic stripe
    else:
        _write_column_lineage_regular(ws_lin, lineage.column_lineage)
        _stripe_rows(ws_lin, ws_lin.max_column)

    _auto_width(ws_lin)

    # ── Sheet 5: Components ─────────────────────────────────────────

    ws_comp = wb.create_sheet("Components")
    comp_hdr = [
        "Name",
        "Type",
        "Description",
        "Input Tables",
        "Output Columns",
        "SQL / Code Snippet",
    ]
    ws_comp.append(comp_hdr)
    _style_header(ws_comp, len(comp_hdr))

    for comp in lineage.components:
        ws_comp.append([
            comp.name,
            comp.component_type,
            comp.description,
            "\n".join(comp.input_tables),
            "\n".join(comp.output_columns),
            comp.sql_text[:200] if comp.sql_text else "",
        ])
    _stripe_rows(ws_comp, len(comp_hdr))
    _auto_width(ws_comp)

    # ── Sheet 6: Data Flow ──────────────────────────────────────────

    ws_flow = wb.create_sheet("Data Flow")
    flow_hdr = ["From", "To", "Columns", "Label"]
    ws_flow.append(flow_hdr)
    _style_header(ws_flow, len(flow_hdr))

    for edge in lineage.data_flow_edges:
        ws_flow.append([
            edge.from_node,
            edge.to_node,
            ", ".join(edge.columns),
            edge.edge_label,
        ])
    _stripe_rows(ws_flow, len(flow_hdr))
    _auto_width(ws_flow)

    # ── Save ────────────────────────────────────────────────────────

    wb.save(str(output_path))
    return output_path
