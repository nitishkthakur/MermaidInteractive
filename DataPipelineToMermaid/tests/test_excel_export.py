"""Unit tests for DataPipelineToMermaid.excel_export."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from DataPipelineToMermaid.excel_export import lineage_to_excel
from DataPipelineToMermaid.models import (
    ColumnInfo,
    ColumnLineage,
    Component,
    DataFlowEdge,
    PipelineLineage,
    TableInfo,
)


# ═══════════════════════════════════════════════════════════════════
# Full fixture export
# ═══════════════════════════════════════════════════════════════════


class TestExcelExportFull:
    """Test Excel export using the full sample_lineage.json fixture."""

    def test_creates_xlsx_file(self, sample_lineage, tmp_output):
        path = lineage_to_excel(sample_lineage, tmp_output / "test.xlsx")
        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_returns_path_object(self, sample_lineage, tmp_output):
        result = lineage_to_excel(sample_lineage, tmp_output / "test.xlsx")
        assert isinstance(result, Path)

    def test_has_six_sheets(self, sample_lineage, tmp_output):
        path = lineage_to_excel(sample_lineage, tmp_output / "test.xlsx")
        wb = load_workbook(path)
        assert wb.sheetnames == [
            "Overview",
            "Source Tables",
            "Target Tables",
            "Column Lineage",
            "Components",
            "Data Flow",
        ]
        wb.close()

    def test_overview_sheet_content(self, sample_lineage, tmp_output):
        path = lineage_to_excel(sample_lineage, tmp_output / "test.xlsx")
        wb = load_workbook(path)
        ws = wb["Overview"]

        # Row 1 = headers
        assert ws.cell(row=1, column=1).value == "Property"
        assert ws.cell(row=1, column=2).value == "Value"

        # Row 2 = pipeline name
        assert ws.cell(row=2, column=1).value == "Pipeline Name"
        assert ws.cell(row=2, column=2).value == "Customer Analytics ETL"

        wb.close()

    def test_source_tables_row_count(self, sample_lineage, tmp_output):
        path = lineage_to_excel(sample_lineage, tmp_output / "test.xlsx")
        wb = load_workbook(path)
        ws = wb["Source Tables"]

        # Count non-header rows (each column of each source table = 1 row)
        total_cols = sum(len(s.columns) for s in sample_lineage.sources)
        # +1 for header row
        assert ws.max_row == total_cols + 1

        wb.close()

    def test_target_tables_row_count(self, sample_lineage, tmp_output):
        path = lineage_to_excel(sample_lineage, tmp_output / "test.xlsx")
        wb = load_workbook(path)
        ws = wb["Target Tables"]

        total_cols = sum(len(t.columns) for t in sample_lineage.targets)
        assert ws.max_row == total_cols + 1

        wb.close()

    def test_column_lineage_row_count(self, sample_lineage, tmp_output):
        path = lineage_to_excel(sample_lineage, tmp_output / "test.xlsx")
        wb = load_workbook(path)
        ws = wb["Column Lineage"]

        assert ws.max_row == len(sample_lineage.column_lineage) + 1

        wb.close()

    def test_column_lineage_headers(self, sample_lineage, tmp_output):
        path = lineage_to_excel(sample_lineage, tmp_output / "test.xlsx")
        wb = load_workbook(path)
        ws = wb["Column Lineage"]

        # Base headers always present (9 cols); Intermediate Steps added as col 9
        # only when at least one row has intermediate_steps populated.
        has_steps = any(cl.intermediate_steps for cl in sample_lineage.column_lineage)
        ncols = 10 if has_steps else 9
        headers = [ws.cell(row=1, column=c).value for c in range(1, ncols + 1)]

        expected_base = [
            "Target Table",
            "Target Column",
            "Source Table(s)",
            "Source Column(s)",
            "Transformation (SQL)",
            "Type",
            "Source File(s)",
            "Target File",
            "Notes",
        ]
        if has_steps:
            expected_base.insert(-1, "Intermediate Steps")

        assert headers == expected_base

        wb.close()

    def test_components_row_count(self, sample_lineage, tmp_output):
        path = lineage_to_excel(sample_lineage, tmp_output / "test.xlsx")
        wb = load_workbook(path)
        ws = wb["Components"]

        assert ws.max_row == len(sample_lineage.components) + 1

        wb.close()

    def test_data_flow_row_count(self, sample_lineage, tmp_output):
        path = lineage_to_excel(sample_lineage, tmp_output / "test.xlsx")
        wb = load_workbook(path)
        ws = wb["Data Flow"]

        assert ws.max_row == len(sample_lineage.data_flow_edges) + 1

        wb.close()

    def test_data_flow_headers(self, sample_lineage, tmp_output):
        path = lineage_to_excel(sample_lineage, tmp_output / "test.xlsx")
        wb = load_workbook(path)
        ws = wb["Data Flow"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert headers == ["From", "To", "Columns", "Label"]

        wb.close()


# ═══════════════════════════════════════════════════════════════════
# Edge cases
# ═══════════════════════════════════════════════════════════════════


class TestExcelExportEdgeCases:
    def test_empty_components_and_edges(self, tmp_output):
        """Pipeline with no components or data_flow_edges."""
        lineage = PipelineLineage(
            pipeline_name="Minimal",
            pipeline_type="sql_query",
            sources=[TableInfo(table_name="src")],
            targets=[TableInfo(table_name="tgt")],
            column_lineage=[
                ColumnLineage(
                    target_table="tgt",
                    target_column="a",
                    source_columns=["src.a"],
                    transformation="a",
                    transformation_type="direct_copy",
                )
            ],
        )
        path = lineage_to_excel(lineage, tmp_output / "minimal.xlsx")
        wb = load_workbook(path)
        assert "Components" in wb.sheetnames
        # Header only for components
        assert wb["Components"].max_row == 1
        assert wb["Data Flow"].max_row == 1
        wb.close()

    def test_tables_without_columns(self, tmp_output):
        """Tables with no column details still produce rows."""
        lineage = PipelineLineage(
            pipeline_name="No Cols",
            pipeline_type="sql_query",
            sources=[TableInfo(table_name="src")],
            targets=[TableInfo(table_name="tgt")],
            column_lineage=[],
        )
        path = lineage_to_excel(lineage, tmp_output / "nocols.xlsx")
        wb = load_workbook(path)
        # Each table without columns → 1 data row (empty columns)
        assert wb["Source Tables"].max_row == 2  # header + 1
        assert wb["Target Tables"].max_row == 2
        wb.close()

    def test_creates_parent_directories(self, tmp_path):
        """Output path with nonexistent parent directories."""
        deep_path = tmp_path / "a" / "b" / "c" / "test.xlsx"
        lineage = PipelineLineage(
            pipeline_name="Deep",
            pipeline_type="sql_query",
            sources=[TableInfo(table_name="s")],
            targets=[TableInfo(table_name="t")],
            column_lineage=[],
        )
        path = lineage_to_excel(lineage, deep_path)
        assert path.exists()

    def test_header_styling_applied(self, sample_lineage, tmp_output):
        """Verify headers have bold white font on blue background."""
        path = lineage_to_excel(sample_lineage, tmp_output / "styled.xlsx")
        wb = load_workbook(path)
        cell = wb["Overview"].cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.fill.start_color.rgb == "FF2F5496" or cell.fill.start_color.rgb == "002F5496"
        wb.close()


# ═══════════════════════════════════════════════════════════════════
# Shared fixture for long / wide mode tests
# ═══════════════════════════════════════════════════════════════════

def _make_multi_hop_lineage() -> PipelineLineage:
    """Build a PipelineLineage with two columns of different chain lengths.

    col_c  — 3-hop chain:  file_a.sql → file_b.mp → file_c.sql
    col_d  — 1-hop chain:  file_a.sql only
    """
    chain_3 = (
        "SOURCES: raw.tbl.raw1, raw.tbl.raw2\n"
        "→ [file_a.sql] raw1 AS col_a\n"
        "→ [file_b.mp] col_a * 2 AS col_b\n"
        "→ [file_c.sql] SUM(col_b) AS col_c"
    )
    chain_1 = (
        "SOURCES: raw.tbl.raw3\n"
        "→ [file_a.sql] raw3 AS col_d"
    )
    return PipelineLineage(
        pipeline_name="Wide/Long Test",
        pipeline_type="sql_query",
        sources=[TableInfo(table_name="raw.tbl")],
        targets=[TableInfo(table_name="final")],
        column_lineage=[
            ColumnLineage(
                target_table="final",
                target_column="col_c",
                source_refs=[
                    {"source_table": "raw.tbl", "source_column": "raw1"},
                    {"source_table": "raw.tbl", "source_column": "raw2"},
                ],
                transformation=chain_3,
                transformation_type="aggregation",
                source_filenames=["file_a.sql", "file_b.mp"],
                filename="file_c.sql",
                notes="test note",
            ),
            ColumnLineage(
                target_table="final",
                target_column="col_d",
                source_refs=[
                    {"source_table": "raw.tbl", "source_column": "raw3"},
                ],
                transformation=chain_1,
                transformation_type="direct_copy",
                source_filenames=[],
                filename="file_a.sql",
                notes="",
            ),
        ],
    )


# ═══════════════════════════════════════════════════════════════════
# Long mode tests
# ═══════════════════════════════════════════════════════════════════


class TestExcelExportLongMode:
    """Tests for artifact_mode='long': one row per file per target column."""

    @pytest.fixture
    def lineage(self):
        return _make_multi_hop_lineage()

    @pytest.fixture
    def wb(self, lineage, tmp_path):
        path = lineage_to_excel(lineage, tmp_path / "long.xlsx", mode="long")
        wb = load_workbook(path)
        yield wb
        wb.close()

    def test_headers_correct(self, wb):
        ws = wb["Column Lineage"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert headers == [
            "Target Table", "Target Column", "Step #", "File",
            "Transformation Step", "Type",
            "Raw Source Tables", "Raw Source Columns", "Notes",
        ]

    def test_row_count_equals_total_hops(self, wb):
        """col_c has 3 hops + col_d has 1 hop = 4 data rows."""
        ws = wb["Column Lineage"]
        assert ws.max_row == 5  # 1 header + 4 data

    def test_key_columns_repeat_every_row(self, wb):
        """target_table and target_column must be non-blank in every data row."""
        ws = wb["Column Lineage"]
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            assert row[0] is not None and row[0] != ""
            assert row[1] is not None and row[1] != ""

    def test_step_numbers_sequential_per_column(self, wb):
        ws = wb["Column Lineage"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        col_c_steps = [r[2] for r in rows if r[1] == "col_c"]
        assert col_c_steps == [1, 2, 3]

    def test_first_step_is_upstream_file(self, wb):
        ws = wb["Column Lineage"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        col_c_files = [r[3] for r in rows if r[1] == "col_c"]
        assert col_c_files[0] == "file_a.sql"
        assert col_c_files[-1] == "file_c.sql"

    def test_transformation_step_populated(self, wb):
        ws = wb["Column Lineage"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert row[4] is not None and row[4] != ""

    def test_short_column_has_one_row(self, wb):
        ws = wb["Column Lineage"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        col_d_rows = [r for r in rows if r[1] == "col_d"]
        assert len(col_d_rows) == 1

    def test_type_repeats_every_row(self, wb):
        """transformation_type should be non-blank in every row."""
        ws = wb["Column Lineage"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert row[5] is not None and row[5] != ""


# ═══════════════════════════════════════════════════════════════════
# Wide mode tests
# ═══════════════════════════════════════════════════════════════════


class TestExcelExportWideMode:
    """Tests for artifact_mode='wide': one row per target column, hops expand rightward."""

    @pytest.fixture
    def lineage(self):
        return _make_multi_hop_lineage()

    @pytest.fixture
    def wb(self, lineage, tmp_path):
        path = lineage_to_excel(lineage, tmp_path / "wide.xlsx", mode="wide")
        wb = load_workbook(path)
        yield wb
        wb.close()

    def test_row_count_equals_target_columns(self, wb):
        """Wide mode: one data row per target column (not per hop)."""
        ws = wb["Column Lineage"]
        assert ws.max_row == 3  # 1 header + 2 target columns

    def test_base_headers(self, wb):
        ws = wb["Column Lineage"]
        assert ws.cell(row=1, column=1).value == "Target Table"
        assert ws.cell(row=1, column=2).value == "Target Column"

    def test_hop_prefix_in_headers(self, wb):
        ws = wb["Column Lineage"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "hop_1_file" in headers
        assert "hop_2_file" in headers
        assert "hop_3_file" in headers

    def test_max_hops_drives_column_count(self, wb):
        """col_c has 3 hops → total cols = 2 base + 3*6 hop = 20."""
        ws = wb["Column Lineage"]
        assert ws.max_column == 2 + 3 * 6  # 20

    def test_hop_n_columns_all_present(self, wb):
        ws = wb["Column Lineage"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for n in range(1, 4):
            for suffix in ("file", "raw_source_tables", "raw_source_columns",
                           "transformation_step", "type", "notes"):
                assert f"hop_{n}_{suffix}" in headers, f"Missing hop_{n}_{suffix}"

    def test_raw_sources_in_hop1_only(self, wb):
        """hop_1_raw_source_tables populated; hop_2+ must be blank."""
        ws = wb["Column Lineage"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        hop1_src_col = headers.index("hop_1_raw_source_tables") + 1
        hop2_src_col = headers.index("hop_2_raw_source_tables") + 1

        # Row 2 = col_c (3 hops, has raw sources)
        assert ws.cell(row=2, column=hop1_src_col).value  # non-blank
        assert not ws.cell(row=2, column=hop2_src_col).value  # blank

    def test_shorter_column_leaves_trailing_hops_blank(self, wb):
        """col_d has 1 hop → hop_2 and hop_3 file cells must be blank."""
        ws = wb["Column Lineage"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        hop2_file_col = headers.index("hop_2_file") + 1
        hop3_file_col = headers.index("hop_3_file") + 1

        # Row 3 = col_d (1 hop only)
        assert not ws.cell(row=3, column=hop2_file_col).value
        assert not ws.cell(row=3, column=hop3_file_col).value

    def test_hop_headers_color_coded(self, wb):
        """Each hop cluster header must have a distinct fill color."""
        ws = wb["Column Lineage"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        hop1_col = headers.index("hop_1_file") + 1
        hop2_col = headers.index("hop_2_file") + 1
        hop3_col = headers.index("hop_3_file") + 1

        c1 = ws.cell(row=1, column=hop1_col).fill.start_color.rgb
        c2 = ws.cell(row=1, column=hop2_col).fill.start_color.rgb
        c3 = ws.cell(row=1, column=hop3_col).fill.start_color.rgb

        # Each hop cluster must have a distinct header color
        assert c1 != c2
        assert c2 != c3
        assert c1 != c3

    def test_type_repeats_in_every_hop(self, wb):
        """hop_n_type must be non-blank for every active hop of a column."""
        ws = wb["Column Lineage"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for n in range(1, 4):  # 3 hops for col_c
            col = headers.index(f"hop_{n}_type") + 1
            assert ws.cell(row=2, column=col).value  # non-blank for col_c

    def test_file_cells_correct_for_3hop_column(self, wb):
        ws = wb["Column Lineage"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        hop1_file = headers.index("hop_1_file") + 1
        hop2_file = headers.index("hop_2_file") + 1
        hop3_file = headers.index("hop_3_file") + 1

        assert ws.cell(row=2, column=hop1_file).value == "file_a.sql"
        assert ws.cell(row=2, column=hop2_file).value == "file_b.mp"
        assert ws.cell(row=2, column=hop3_file).value == "file_c.sql"
